"""
GNEM document relevance pipeline.

The original 3-stage metadata/PDF/first-page flow is still the entry point, but
the pipeline now extends into a phase-1 RAG corpus construction workflow:
  - metadata scoring
  - document-card generation
  - heuristic + embedding filtering
  - rubric classifier reranking
  - shortlist-only LLM judging

Designed for transparent, review-ready filtering with auditable evidence.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
import shutil
import sqlite3
import textwrap
import time
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

import pandas as pd
import requests
from dotenv import load_dotenv
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover - optional formatting dependency
    load_workbook = None
    Alignment = None
    Font = None
    PatternFill = None
    get_column_letter = None
try:
    from .gnem_rag_helpers import (
        apply_diversity_pass,
        assess_document_credibility,
        build_document_card,
        build_grounding_dictionaries,
        build_local_document_index,
        build_local_text_index,
        classify_document_card,
        detect_grounding_file,
        final_decision_reason,
        grounding_summary_payload,
        probable_document,
        resolve_document_paths,
        score_document_card,
    )
    from .embedding_runtime import EmbeddingConfig, EmbeddingRuntime
except ImportError:  # pragma: no cover - allows direct file execution from repo root
    from tavily_ev_automation.gnem_rag_helpers import (
        apply_diversity_pass,
        assess_document_credibility,
        build_document_card,
        build_grounding_dictionaries,
        build_local_document_index,
        build_local_text_index,
        classify_document_card,
        detect_grounding_file,
        final_decision_reason,
        grounding_summary_payload,
        probable_document,
        resolve_document_paths,
        score_document_card,
    )
    from tavily_ev_automation.embedding_runtime import EmbeddingConfig, EmbeddingRuntime
try:
    from tavily import TavilyClient
except Exception:  # pragma: no cover - handled at runtime
    TavilyClient = None

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover - handled at runtime
    PdfReader = None
try:
    import fitz  # PyMuPDF
except Exception:  # pragma: no cover - handled at runtime
    fitz = None

def load_shared_env_files() -> None:
    repo_root = Path(__file__).resolve().parent.parent
    for env_path in [
        repo_root / ".env",
        repo_root / "evAutomationUpdated" / ".env",
    ]:
        if env_path.exists():
            load_dotenv(env_path, override=False)
    load_dotenv(override=False)


def disable_broken_local_proxies() -> None:
    broken_markers = ("127.0.0.1:9", "localhost:9")
    for key in ("HTTP_PROXY", "HTTPS_PROXY", "ALL_PROXY", "http_proxy", "https_proxy", "all_proxy"):
        value = (os.environ.get(key) or "").strip().lower()
        if value and any(marker in value for marker in broken_markers):
            os.environ.pop(key, None)


load_shared_env_files()
disable_broken_local_proxies()

ILLEGAL_XLSX_CHARS_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")
DEFAULT_QUERY_FILE = Path("data") / "queries" / "queries_1000.txt"
LEGACY_QUERY_FILE = Path("queries_1000.txt")
DEFAULT_GROUNDING_XLSX = Path("data") / "grounding" / "GA_Automotive Landscape_All_Companies (1).xlsx"
LEGACY_GROUNDING_XLSX = Path("GA_Automotive Landscape_All_Companies (1).xlsx")
ALT_LEGACY_GROUNDING_XLSX = Path("GA_Automotive Landscape_All_Companies.xlsx")
DEFAULT_GROUNDING_GEOJSON = Path("data") / "grounding" / "Counties_Georgia.geojson"
LEGACY_GROUNDING_GEOJSON = Path("Counties_Georgia.geojson")
DEFAULT_GROUNDING_DOCX = Path("data") / "grounding" / "GNEM Supply Chain.docx"
LEGACY_GROUNDING_DOCX = Path("GNEM Supply Chain.docx")
DEFAULT_LOCAL_DOC_DIR = Path("data") / "corpus" / "documents"
LEGACY_LOCAL_DOC_DIR = Path("corpus_gnem_battery")
DEFAULT_LOCAL_TEXT_DIR = Path("data") / "corpus" / "text"
LEGACY_LOCAL_TEXT_DIR = Path("corpus_text")
DEFAULT_OUTPUTS_DIR = Path("outputs") / "pipeline_runs"
TRACKING_QUERY_PARAMS = {
    "fbclid",
    "gclid",
    "mc_cid",
    "mc_eid",
    "mkt_tok",
    "ref",
    "spm",
    "trk",
    "utm_campaign",
    "utm_content",
    "utm_id",
    "utm_medium",
    "utm_source",
    "utm_term",
}


GOLDEN_SUMMARY_DEFAULT = """
GNEM's Cognitive Supply Chain Digital Twin needs documents that build a living,
risk-aware supply chain network for the U.S. EV battery ecosystem, prioritizing
the Battery Belt (Georgia + Southeast). National/global documents are relevant
only if they can be mapped to Georgia/Southeast nodes, facilities, suppliers,
logistics routes, or localization decisions.

The goal is to support a Resilience Control Tower that can map multi-tier
relationships, quantify dependencies and bottlenecks, simulate disruption risk
under uncertainty, and prioritize localization actions for OEMs and state/industry
leaders.

A document is relevant if it contains at least two digital-twin signals:
- Network structure (tiering, supplier-customer links, ecosystem maps)
- Node attributes (locations, capacities, ownership, processes, timelines)
- Edge attributes (material flows, logistics routes, lead times, import/export exposure)
- Risk variables (concentration, permitting/regulatory/trade exposure, disruptions)
- Optimization levers (gap analysis, incentives, workforce/site constraints)
- Scenario modeling (resilience scoring, what-if stress tests, simulation)

It is also relevant if it clearly answers at least one key filter question with
specific entities, locations, capacity, or timing:
who makes what where at what scale; what relationships/flows exist; what
constraints/risks apply; what supports localization decisions.

Prioritize:
- Value-chain mapping (minerals -> refining -> CAM/anode -> cell -> pack -> OEM -> recycling)
- Georgia/Southeast ecosystem evidence (facilities, investments, suppliers, tiering)
- Trade/logistics dependencies (Savannah/Brunswick ports, rail, hazmat)
- Policy and incentives (IRA, state programs, permitting, workforce)
- Resilience and bottleneck analyses
- Government/research/policy reports and credible company filings/announcements

De-prioritize generic EV news without supply-chain specifics, chemistry-only
papers without ecosystem relevance, unverifiable marketing pages, and duplicates.
""".strip()


DIGITAL_TWIN_SIGNAL_KEYWORDS: dict[str, list[str]] = {
    "network_structure": [
        "tier 1",
        "tier 2",
        "tiering",
        "supplier",
        "customer",
        "supply chain map",
        "ecosystem map",
        "value chain",
        "network",
        "oem",
    ],
    "node_attributes": [
        "facility",
        "plant",
        "location",
        "capacity",
        "gwh",
        "mwh",
        "tons per year",
        "ownership",
        "commissioning",
        "timeline",
        "expansion",
        "site",
    ],
    "edge_attributes": [
        "material flow",
        "logistics",
        "route",
        "lead time",
        "import",
        "export",
        "shipping",
        "freight",
        "port",
        "rail",
        "hazmat",
    ],
    "risk_variables": [
        "risk",
        "bottleneck",
        "concentration",
        "single source",
        "trade exposure",
        "regulatory",
        "permitting",
        "disruption",
        "dependency",
        "shortage",
    ],
    "optimization_levers": [
        "localization",
        "gap analysis",
        "incentive",
        "tax credit",
        "workforce",
        "site readiness",
        "site selection",
        "grant",
        "capex",
        "investment",
    ],
    "scenario_modeling": [
        "scenario",
        "what-if",
        "stress test",
        "simulation",
        "resilience score",
        "digital twin",
        "control tower",
        "sensitivity analysis",
    ],
}


VALUE_CHAIN_KEYWORDS = [
    "minerals",
    "mining",
    "refining",
    "cathode",
    "cam",
    "anode",
    "separator",
    "electrolyte",
    "cell",
    "module",
    "pack",
    "oem",
    "recycling",
    "black mass",
    "hydrometallurgy",
]

REGION_KEYWORDS = [
    "georgia",
    "southeast",
    "battery belt",
    "atlanta",
    "savannah",
    "brunswick",
    "tennessee",
    "alabama",
    "south carolina",
    "north carolina",
    "kentucky",
    "mississippi",
]

POLICY_LOGISTICS_KEYWORDS = [
    "ira",
    "inflation reduction act",
    "doe",
    "state incentive",
    "tax credit",
    "permitting",
    "workforce",
    "savannah port",
    "brunswick port",
    "rail",
    "hazmat",
]

NEGATIVE_KEYWORDS = [
    "buy now",
    "shop now",
    "affiliate",
    "sponsored",
    "press release wire",
    "click here",
    "lifestyle",
    "celebrity",
    "top 10",
    "best ev",
    "review",
    "coupon",
    "deal",
    "sponsored content",
    "native advertising",
    "news roundup",
    "price prediction",
    "crypto",
    "betting",
    "giveaway",
    "chemistry only",
]

PDF_HINT_KEYWORDS = [
    "report",
    "white paper",
    "brief",
    "fact sheet",
    "study",
    "roadmap",
    "slides",
    "presentation",
    "publication",
    "year-end",
    "outlook",
]

DOCUMENT_QUALITY_KEYWORDS = [
    "report",
    "white paper",
    "roadmap",
    "policy brief",
    "analysis",
    "study",
    "annual report",
    "investor presentation",
    "10-k",
    "8-k",
    "fact sheet",
    "directory",
    "dataset",
    "filing",
]

BLOCKLIST_SOURCE_DOMAIN_HINTS = {
    "ainvest.com",
}

LOW_TRUST_FINANCE_NEWS_DOMAIN_HINTS = {
    "ainvest.com",
    "benzinga.com",
    "investing.com",
    "investorplace.com",
    "marketbeat.com",
    "marketscreener.com",
    "fool.com",
    "247wallst.com",
    "seekingalpha.com",
    "stocktwits.com",
    "tipranks.com",
    "thestreet.com",
    "zacks.com",
}

GEORGIA_NODE_KEYWORDS = [
    "georgia",
    "southeast",
    "battery belt",
    "atlanta",
    "savannah",
    "brunswick",
    "chatham county",
    "bartow county",
    "coweta county",
    "jackson county",
    "west point",
    "commerce ga",
]

NATIONAL_GLOBAL_KEYWORDS = [
    "united states",
    "u.s.",
    "us ",
    "north america",
    "global",
    "international",
    "worldwide",
]

QUERY_ENHANCEMENT_HINTS = [
    "supplier tier network",
    "facility capacity timeline",
    "Georgia Southeast",
    "risk resilience localization",
]


@dataclass
class ScoreBreakdown:
    metadata_score: float
    category_hit_count: int
    category_hits: str
    signal_keyword_hits: int
    region_hit: int
    value_chain_hits: int
    policy_logistics_hits: int
    question_coverage: int
    georgia_node_hits: int
    specificity_hits: int
    document_quality_hits: int
    two_signal_rule: int
    question_specifics_rule: int
    source_credibility: float
    penalty_points: float
    lexical_similarity: float


@dataclass
class LLMConfig:
    provider: str
    model: str
    base_url: str
    api_key: str
    timeout_sec: int
    temperature: float
    max_text_chars: int
    first_page_weight: float
    retries: int = 2
    retry_sleep_sec: float = 1.0


def normalize_space(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "")).strip()


def as_text(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return normalize_space(str(value))


def sanitize_filename(name: str) -> str:
    name = normalize_space(name)
    name = re.sub(r'[<>:"/\\|?*]', "_", name)
    name = re.sub(r"\s+", "_", name)
    name = re.sub(r"_+", "_", name)
    return name[:180] if len(name) > 180 else name


def safe_output_filename(directory: Path, name: str, *, max_full_path: int = 240) -> str:
    safe_name = sanitize_filename(name) or "document"
    suffix = Path(safe_name).suffix
    stem = Path(safe_name).stem if suffix else safe_name
    try:
        dir_len = len(str(directory.resolve()))
    except Exception:
        dir_len = len(str(directory))
    max_name_len = max(48, max_full_path - dir_len - 1)
    if len(safe_name) <= max_name_len:
        return safe_name
    allowed_stem_len = max(16, max_name_len - len(suffix))
    trimmed_stem = stem[:allowed_stem_len].rstrip("._- ")
    return f"{trimmed_stem}{suffix}" if suffix else trimmed_stem


def filename_to_text(name: str) -> str:
    txt = requests.utils.unquote(as_text(name))
    txt = re.sub(r"\.[a-z0-9]{2,5}$", "", txt, flags=re.IGNORECASE)
    txt = re.sub(r"_DOC_\d+(?:_\d+)?$", "", txt, flags=re.IGNORECASE)
    txt = re.sub(r"[_\-]+", " ", txt)
    return normalize_space(txt)


def canonical_filename_key(name: str) -> str:
    return filename_to_text(name).lower()


def get_domain(url: str) -> str:
    return (urlparse(url).netloc or "").lower()


def is_pdf_url(url: str) -> bool:
    path = urlparse(url).path.lower()
    return path.endswith(".pdf")


def token_counter(text: str) -> Counter[str]:
    tokens = re.findall(r"[a-z0-9][a-z0-9\-\+\.]*", text.lower())
    return Counter(tokens)


def cosine_similarity_text(a: str, b: str) -> float:
    ca = token_counter(a)
    cb = token_counter(b)
    if not ca or not cb:
        return 0.0
    all_terms = set(ca) | set(cb)
    dot = sum(ca[t] * cb[t] for t in all_terms)
    na = math.sqrt(sum(v * v for v in ca.values()))
    nb = math.sqrt(sum(v * v for v in cb.values()))
    if na == 0 or nb == 0:
        return 0.0
    return max(0.0, min(1.0, dot / (na * nb)))


def keyword_hits(text: str, keywords: Iterable[str]) -> tuple[int, list[str]]:
    txt = text.lower()
    hits = [kw for kw in keywords if kw in txt]
    return len(hits), hits


def category_hits(text: str) -> tuple[int, int, list[str]]:
    txt = text.lower()
    cat_names: list[str] = []
    total_hits = 0
    for cat, kws in DIGITAL_TWIN_SIGNAL_KEYWORDS.items():
        cat_hit = False
        for kw in kws:
            if kw in txt:
                cat_hit = True
                total_hits += 1
        if cat_hit:
            cat_names.append(cat)
    return len(cat_names), total_hits, cat_names


def source_credibility_score(domain: str) -> float:
    if not domain:
        return 3.0
    d = domain.lower()
    if any(d == hint or d.endswith(f".{hint}") for hint in BLOCKLIST_SOURCE_DOMAIN_HINTS):
        return 0.5
    if any(d == hint or d.endswith(f".{hint}") for hint in LOW_TRUST_FINANCE_NEWS_DOMAIN_HINTS):
        return 1.5
    if d.endswith(".gov"):
        return 15.0
    if d.endswith(".edu"):
        return 13.0
    if d.endswith(".org"):
        return 10.0
    if any(x in d for x in ["sec.gov", "doe.gov", "energy.gov", "nist.gov", "epa.gov"]):
        return 15.0
    if any(x in d for x in ["reuters", "bloomberg", "wsj", "ft.com"]):
        return 8.0
    if any(x in d for x in ["linkedin", "facebook", "instagram", "x.com", "twitter"]):
        return 2.0
    return 6.0


def question_coverage_count(text: str) -> int:
    txt = text.lower()
    groups = [
        ["who", "supplier", "manufacturer", "company", "producer"],
        ["what", "product", "material", "cell", "module", "pack", "cam", "anode"],
        ["where", "georgia", "southeast", "plant", "facility", "location"],
        ["scale", "capacity", "gwh", "mwh", "tons", "mtpa"],
        ["relationship", "customer", "offtake", "agreement", "tier", "flow"],
        ["constraints", "risk", "permitting", "trade", "lead time", "bottleneck"],
        ["localization", "incentive", "workforce", "site readiness", "control tower"],
    ]
    count = 0
    for kws in groups:
        if any(kw in txt for kw in kws):
            count += 1
    return count


def specificity_signal_count(text: str) -> int:
    txt = text.lower()
    rules = [
        r"\b\d+(?:\.\d+)?\s?(gwh|mwh|kwh|mtpa|tons|tonnes|million|billion|%)\b",
        r"\b20\d{2}\b",
        r"\b(inc|llc|corp|corporation|co\.|ltd|plc|ag)\b",
        r"\b(offtake|joint venture|jv|supplier agreement|customer agreement)\b",
        r"\b(plant|facility|site|county|port|rail)\b",
    ]
    return sum(1 for pat in rules if re.search(pat, txt))


def strip_query_pdf_hint(query: str) -> str:
    return normalize_space(re.sub(r"\bfiletype\s*:\s*pdf\b", " ", query or "", flags=re.I))


def canonicalize_url(url: str) -> str:
    raw = normalize_space(url)
    if not raw:
        return ""
    parsed = urlparse(raw)
    scheme = (parsed.scheme or "https").lower()
    netloc = (parsed.netloc or "").lower()
    path = parsed.path or "/"
    if path != "/" and path.endswith("/"):
        path = path.rstrip("/")
    cleaned_query = [
        (key, value)
        for key, value in parse_qsl(parsed.query, keep_blank_values=False)
        if key.lower() not in TRACKING_QUERY_PARAMS and not key.lower().startswith("utm_")
    ]
    return urlunparse((scheme, netloc, path, "", urlencode(cleaned_query, doseq=True), ""))


def enhanced_query_for_tavily(query: str, *, include_pdf_hint: bool = False) -> str:
    q = strip_query_pdf_hint(query)
    ql = q.lower()
    additions: list[str] = []
    if include_pdf_hint and "filetype:pdf" not in ql:
        additions.append("filetype:pdf")
    for hint in QUERY_ENHANCEMENT_HINTS:
        if hint.lower() not in ql:
            additions.append(hint)
    return normalize_space(f"{q} {' '.join(additions)}")[:420]


def query_variants_for_mode(query: str, *, query_mode: str, query_enhancement: bool) -> list[tuple[str, str]]:
    base_query = normalize_space(query)
    variants: list[tuple[str, str]] = []
    if query_mode == "pdf_only":
        variants.append(
            (
                "pdf_only",
                enhanced_query_for_tavily(base_query, include_pdf_hint=True)
                if query_enhancement else normalize_space(f"{strip_query_pdf_hint(base_query)} filetype:pdf"),
            )
        )
    elif query_mode == "web_only":
        variants.append(
            (
                "web_only",
                enhanced_query_for_tavily(base_query, include_pdf_hint=False)
                if query_enhancement else strip_query_pdf_hint(base_query),
            )
        )
    elif query_mode == "hybrid":
        variants.extend(
            [
                (
                    "pdf_only",
                    enhanced_query_for_tavily(base_query, include_pdf_hint=True)
                    if query_enhancement else normalize_space(f"{strip_query_pdf_hint(base_query)} filetype:pdf"),
                ),
                (
                    "web_only",
                    enhanced_query_for_tavily(base_query, include_pdf_hint=False)
                    if query_enhancement else strip_query_pdf_hint(base_query),
                ),
            ]
        )
    else:
        variants.append(("as_is", base_query))

    unique_variants: list[tuple[str, str]] = []
    seen: set[str] = set()
    for label, value in variants:
        cleaned = normalize_space(value)
        if not cleaned:
            continue
        key = cleaned.lower()
        if key in seen:
            continue
        seen.add(key)
        unique_variants.append((label, cleaned))
    return unique_variants


def clamp_score(value: float, low: float = 0.0, high: float = 100.0) -> float:
    return max(low, min(high, float(value)))


def to_int_01(value: object) -> int:
    if isinstance(value, bool):
        return 1 if value else 0
    try:
        return 1 if int(float(value)) > 0 else 0
    except Exception:
        return 0


def extract_json_object(text: str) -> dict[str, Any] | None:
    txt = normalize_space(text)
    if not txt:
        return None
    try:
        loaded = json.loads(txt)
        return loaded if isinstance(loaded, dict) else None
    except Exception:
        pass

    block_match = re.search(r"\{.*\}", text, flags=re.DOTALL)
    if not block_match:
        return None
    block = block_match.group(0).strip()
    try:
        loaded = json.loads(block)
        return loaded if isinstance(loaded, dict) else None
    except Exception:
        return None


def llm_chat_json(
    *,
    cfg: LLMConfig,
    system_prompt: str,
    user_prompt: str,
) -> tuple[dict[str, Any] | None, str]:
    provider = (cfg.provider or "").strip().lower()
    last_error = ""
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ]

    for attempt in range(1, cfg.retries + 2):
        try:
            if provider == "ollama":
                # Native Ollama chat endpoint.
                native_url = cfg.base_url.rstrip("/") + "/api/chat"
                native_payload = {
                    "model": cfg.model,
                    "messages": messages,
                    "stream": False,
                    "format": "json",
                    "options": {"temperature": cfg.temperature},
                }
                try:
                    resp = requests.post(native_url, json=native_payload, timeout=cfg.timeout_sec)
                    resp.raise_for_status()
                    data = resp.json()
                    content = as_text((data.get("message") or {}).get("content"))
                    parsed = extract_json_object(content)
                    if parsed is not None:
                        return parsed, "ok"
                    last_error = "invalid_json_from_ollama_native"
                except requests.HTTPError as exc:
                    # Older/newer builds can expose only OpenAI-compatible routes.
                    if exc.response is None or exc.response.status_code != 404:
                        raise

                # OpenAI-compatible Ollama endpoint fallback.
                compat_url = cfg.base_url.rstrip("/") + "/v1/chat/completions"
                compat_payload = {
                    "model": cfg.model,
                    "messages": messages,
                    "temperature": cfg.temperature,
                    "response_format": {"type": "json_object"},
                }
                resp2 = requests.post(compat_url, json=compat_payload, timeout=cfg.timeout_sec)
                resp2.raise_for_status()
                data2 = resp2.json()
                content2 = as_text((((data2.get("choices") or [{}])[0].get("message") or {}).get("content")))
                parsed2 = extract_json_object(content2)
                if parsed2 is not None:
                    return parsed2, "ok"
                last_error = "invalid_json_from_ollama_compat"

            elif provider == "openai":
                if not cfg.api_key:
                    return None, "missing_openai_api_key"
                url = cfg.base_url.rstrip("/") + "/v1/chat/completions"
                headers = {
                    "Authorization": f"Bearer {cfg.api_key}",
                    "Content-Type": "application/json",
                }
                payload = {
                    "model": cfg.model,
                    "messages": messages,
                    "temperature": cfg.temperature,
                    "response_format": {"type": "json_object"},
                }
                resp = requests.post(url, headers=headers, json=payload, timeout=cfg.timeout_sec)
                resp.raise_for_status()
                data = resp.json()
                content = as_text((((data.get("choices") or [{}])[0].get("message") or {}).get("content")))
                parsed = extract_json_object(content)
                if parsed is not None:
                    return parsed, "ok"
                last_error = "invalid_json_from_openai"
            else:
                return None, f"unsupported_provider:{provider}"

        except Exception as exc:
            last_error = f"llm_call_failed:{exc}"

        if attempt < cfg.retries + 1:
            time.sleep(cfg.retry_sleep_sec)

    return None, last_error or "llm_unknown_error"


def llm_first_page_assessment(
    *,
    cfg: LLMConfig,
    title: str,
    source_domain: str,
    metadata_score: float,
    golden_summary: str,
    first_page_text: str,
) -> dict[str, Any]:
    first_page_text = normalize_space(first_page_text)[: cfg.max_text_chars]
    if not first_page_text:
        return {"status": "empty_first_page"}

    system_prompt = (
        "You are an expert supply-chain relevance evaluator for GNEM's EV battery digital twin. "
        "Return strict JSON only. No markdown."
    )
    user_prompt = textwrap.dedent(
        f"""
        Evaluate the PDF first page for relevance to the GNEM use case below.

        GNEM_GOLDEN_SUMMARY:
        {golden_summary}

        CANDIDATE_METADATA:
        - title: {title}
        - source_domain: {source_domain}
        - metadata_score: {metadata_score}

        FIRST_PAGE_TEXT:
        {first_page_text}

        Return a JSON object with exactly these keys:
        {{
          "summary": "max 120 words",
          "relevance_score": 0-100,
          "digital_twin_signal_count": 0-6,
          "key_signals": ["..."],
          "answers_key_question_with_specifics": 0 or 1,
          "entity_specificity_count": 0-20,
          "georgia_southeast_anchor": 0 or 1,
          "recommendation": "keep" or "drop",
          "confidence": 0-1,
          "rationale": "max 60 words"
        }}
        """
    ).strip()

    data, status = llm_chat_json(cfg=cfg, system_prompt=system_prompt, user_prompt=user_prompt)
    if not data:
        return {"status": status}

    summary = as_text(data.get("summary"))[:900]
    if not summary:
        summary = first_page_text[:750]
    key_signals_raw = data.get("key_signals")
    if isinstance(key_signals_raw, list):
        key_signals = [as_text(x) for x in key_signals_raw if as_text(x)]
    else:
        key_signals = []
    try:
        confidence = float(data.get("confidence", 0.0))
    except Exception:
        confidence = 0.0

    return {
        "status": "ok",
        "summary": summary,
        "relevance_score": round(clamp_score(float(data.get("relevance_score", 0.0))), 2),
        "digital_twin_signal_count": max(0, min(6, int(float(data.get("digital_twin_signal_count", 0))))),
        "key_signals": key_signals,
        "answers_key_question_with_specifics": to_int_01(data.get("answers_key_question_with_specifics", 0)),
        "entity_specificity_count": max(0, min(20, int(float(data.get("entity_specificity_count", 0))))),
        "georgia_southeast_anchor": to_int_01(data.get("georgia_southeast_anchor", 0)),
        "recommendation": "keep" if as_text(data.get("recommendation")).lower() == "keep" else "drop",
        "confidence": round(max(0.0, min(1.0, confidence)), 3),
        "rationale": as_text(data.get("rationale"))[:400],
    }


def metadata_score_for_candidate(
    *,
    title: str,
    content: str,
    description: str,
    query: str,
    url: str,
    tavily_score_raw: float,
    golden_summary: str,
) -> ScoreBreakdown:
    text = normalize_space(" ".join([title, content, description, query, url]))
    domain = get_domain(url)

    cat_count, signal_hits_total, cat_names = category_hits(text)
    vc_hits, _ = keyword_hits(text, VALUE_CHAIN_KEYWORDS)
    region_hits, _ = keyword_hits(text, REGION_KEYWORDS)
    georgia_node_hits, _ = keyword_hits(text, GEORGIA_NODE_KEYWORDS)
    pl_hits, _ = keyword_hits(text, POLICY_LOGISTICS_KEYWORDS)
    doc_quality_hits, _ = keyword_hits(text, DOCUMENT_QUALITY_KEYWORDS)
    question_hits = question_coverage_count(text)
    specificity_hits = specificity_signal_count(text)
    neg_hits, _ = keyword_hits(text, NEGATIVE_KEYWORDS)
    national_global_hits, _ = keyword_hits(text, NATIONAL_GLOBAL_KEYWORDS)
    lexical_sim = cosine_similarity_text(text, golden_summary) * 100.0

    tavily_score_pct = max(0.0, min(100.0, float(tavily_score_raw or 0.0) * 100.0))
    cat_score = min(100.0, (cat_count / 6.0) * 100.0)
    value_chain_score = min(100.0, (vc_hits / 8.0) * 100.0)
    region_score = 100.0 if region_hits > 0 else 0.0
    georgia_node_score = min(100.0, (georgia_node_hits / 6.0) * 100.0)
    policy_score = min(100.0, (pl_hits / 5.0) * 100.0)
    question_score = min(100.0, (question_hits / 7.0) * 100.0)
    specificity_score = min(100.0, (specificity_hits / 5.0) * 100.0)
    doc_quality_score = min(100.0, (doc_quality_hits / 6.0) * 100.0)
    two_signal_rule = 1 if cat_count >= 2 else 0
    question_specifics_rule = 1 if question_hits >= 1 and specificity_hits >= 1 else 0
    cred_score = source_credibility_score(domain)
    penalty = min(24.0, float(neg_hits * 4))
    # Penalize national/global docs that do not map to Georgia/Southeast evidence.
    if national_global_hits > 0 and georgia_node_hits == 0 and region_hits == 0:
        penalty += 8.0
    rule_bonus = 5.0 if two_signal_rule else 0.0
    if question_specifics_rule:
        rule_bonus += 3.0

    meta_score = (
        0.12 * tavily_score_pct
        + 0.24 * cat_score
        + 0.12 * value_chain_score
        + 0.12 * region_score
        + 0.08 * georgia_node_score
        + 0.08 * policy_score
        + 0.08 * question_score
        + 0.06 * specificity_score
        + 0.06 * doc_quality_score
        + 0.10 * lexical_sim
        + cred_score
        + rule_bonus
        - penalty
    )
    meta_score = max(0.0, min(100.0, meta_score))

    return ScoreBreakdown(
        metadata_score=round(meta_score, 2),
        category_hit_count=cat_count,
        category_hits=", ".join(cat_names),
        signal_keyword_hits=signal_hits_total,
        region_hit=1 if region_hits > 0 else 0,
        value_chain_hits=vc_hits,
        policy_logistics_hits=pl_hits,
        question_coverage=question_hits,
        georgia_node_hits=georgia_node_hits,
        specificity_hits=specificity_hits,
        document_quality_hits=doc_quality_hits,
        two_signal_rule=two_signal_rule,
        question_specifics_rule=question_specifics_rule,
        source_credibility=round(cred_score, 2),
        penalty_points=round(penalty, 2),
        lexical_similarity=round(lexical_sim, 2),
    )


def first_page_summary(text: str, max_sentences: int = 4, max_chars: int = 750) -> str:
    txt = normalize_space(text)
    if not txt:
        return ""
    sentences = re.split(r"(?<=[\.\!\?])\s+", txt)
    if len(sentences) <= max_sentences:
        return txt[:max_chars]

    keyword_pool = (
        VALUE_CHAIN_KEYWORDS
        + REGION_KEYWORDS
        + POLICY_LOGISTICS_KEYWORDS
        + [kw for kws in DIGITAL_TWIN_SIGNAL_KEYWORDS.values() for kw in kws]
    )
    scored: list[tuple[int, float, str]] = []
    for idx, s in enumerate(sentences):
        s_lower = s.lower()
        score = sum(1.0 for kw in keyword_pool if kw in s_lower)
        score += 0.6 * len(re.findall(r"\b\d+(?:\.\d+)?\b", s_lower))
        score += 0.6 * specificity_signal_count(s_lower)
        if "georgia" in s_lower or "southeast" in s_lower:
            score += 1.0
        scored.append((idx, score, s))

    picked_indices: set[int] = {0}
    for idx, _, _ in sorted(scored, key=lambda x: x[1], reverse=True):
        picked_indices.add(idx)
        if len(picked_indices) >= max_sentences:
            break
    picked_sorted = [sentences[i] for i in sorted(picked_indices) if i < len(sentences)]
    summary = " ".join(picked_sorted)
    return summary[:max_chars]


def first_page_score(
    *,
    first_page_text: str,
    first_page_summary_text: str,
    golden_summary: str,
    metadata_score: float,
) -> tuple[float, dict[str, float | int | str]]:
    fp_text = normalize_space(first_page_text)
    fp_summary = normalize_space(first_page_summary_text)
    combined = normalize_space(f"{fp_summary} {fp_text}")

    cat_count, signal_hits_total, cat_names = category_hits(combined)
    vc_hits, _ = keyword_hits(combined, VALUE_CHAIN_KEYWORDS)
    region_hits, _ = keyword_hits(combined, REGION_KEYWORDS)
    georgia_node_hits, _ = keyword_hits(combined, GEORGIA_NODE_KEYWORDS)
    pl_hits, _ = keyword_hits(combined, POLICY_LOGISTICS_KEYWORDS)
    doc_quality_hits, _ = keyword_hits(combined, DOCUMENT_QUALITY_KEYWORDS)
    question_hits = question_coverage_count(combined)
    specificity_hits = specificity_signal_count(combined)
    lexical_sim_full = cosine_similarity_text(combined, golden_summary) * 100.0
    lexical_sim_summary = cosine_similarity_text(fp_summary, golden_summary) * 100.0
    lexical_sim = 0.7 * lexical_sim_full + 0.3 * lexical_sim_summary
    entity_count = len(
        re.findall(r"\b\d+(?:\.\d+)?\s?(gwh|mwh|kwh|mtpa|tons|million|billion|%)?\b", combined.lower())
    )
    neg_hits, _ = keyword_hits(combined, NEGATIVE_KEYWORDS)

    cat_score = min(100.0, (cat_count / 6.0) * 100.0)
    vc_score = min(100.0, (vc_hits / 8.0) * 100.0)
    region_score = 100.0 if region_hits > 0 else 0.0
    georgia_node_score = min(100.0, (georgia_node_hits / 6.0) * 100.0)
    policy_score = min(100.0, (pl_hits / 5.0) * 100.0)
    question_score = min(100.0, (question_hits / 7.0) * 100.0)
    entity_score = min(100.0, entity_count * 12.5)
    specificity_score = min(100.0, (specificity_hits / 5.0) * 100.0)
    doc_quality_score = min(100.0, (doc_quality_hits / 6.0) * 100.0)
    two_signal_rule = 1 if cat_count >= 2 else 0
    question_specifics_rule = 1 if question_hits >= 1 and entity_count >= 1 else 0
    rule_bonus = (5.0 if two_signal_rule else 0.0) + (3.0 if question_specifics_rule else 0.0)
    if len(fp_text) >= 350:
        length_adjust = 4.0
    elif len(fp_text) >= 180:
        length_adjust = 1.0
    elif len(fp_text) < 90:
        length_adjust = -6.0
    else:
        length_adjust = -2.0
    penalty = min(12.0, float(neg_hits * 4))

    score = (
        0.26 * lexical_sim
        + 0.22 * cat_score
        + 0.12 * vc_score
        + 0.08 * region_score
        + 0.08 * georgia_node_score
        + 0.08 * policy_score
        + 0.08 * question_score
        + 0.08 * specificity_score
        + 0.06 * doc_quality_score
        + 0.06 * entity_score
        + 0.04 * float(metadata_score)
        + rule_bonus
        + length_adjust
        - penalty
    )
    score = max(0.0, min(100.0, score))

    details = {
        "FirstPage_Category_Count": cat_count,
        "FirstPage_Category_Hits": ", ".join(cat_names),
        "FirstPage_Signal_Hits": signal_hits_total,
        "FirstPage_ValueChain_Hits": vc_hits,
        "FirstPage_Region_Hit": 1 if region_hits > 0 else 0,
        "FirstPage_GeorgiaNode_Hits": georgia_node_hits,
        "FirstPage_PolicyLogistics_Hits": pl_hits,
        "FirstPage_Question_Coverage": question_hits,
        "FirstPage_Specificity_Hits": specificity_hits,
        "FirstPage_DocQuality_Hits": doc_quality_hits,
        "FirstPage_Entity_Count": entity_count,
        "FirstPage_TwoSignal_Rule": two_signal_rule,
        "FirstPage_QuestionSpecifics_Rule": question_specifics_rule,
        "FirstPage_Lexical_Similarity_Full": round(lexical_sim_full, 2),
        "FirstPage_Lexical_Similarity_Summary": round(lexical_sim_summary, 2),
        "FirstPage_Lexical_Similarity": round(lexical_sim, 2),
        "FirstPage_Penalty_Points": round(penalty, 2),
        "FirstPage_Score": round(score, 2),
        "Metadata_Score_Input": round(float(metadata_score), 2),
    }
    return round(score, 2), details


def blended_first_page_outcome(
    *,
    rule_score: float,
    rule_details: dict[str, float | int | str],
    llm_result: dict[str, Any] | None,
    llm_weight: float,
) -> tuple[float, dict[str, Any]]:
    out: dict[str, Any] = {}
    out["FirstPage_Score_Rule"] = round(float(rule_score), 2)
    out["FirstPage_Score_LLM"] = ""
    out["FirstPage_Score_Blend_Weight_LLM"] = round(float(llm_weight), 3)
    out["FirstPage_Used_LLM"] = 0
    out["FirstPage_LLM_Status"] = "not_used"
    out["FirstPage_LLM_DigitalTwinSignal_Count"] = 0
    out["FirstPage_LLM_QuestionSpecifics_Rule"] = 0
    out["FirstPage_LLM_Georgia_Anchor"] = 0
    out["FirstPage_LLM_EntitySpecificity_Count"] = 0
    out["FirstPage_LLM_Recommendation"] = ""
    out["FirstPage_LLM_Confidence"] = 0.0
    out["FirstPage_LLM_KeySignals"] = ""
    out["FirstPage_LLM_Rationale"] = ""
    out["FirstPage_LLM_Rule_Pass"] = 0
    out["FirstPage_LLM_Summary"] = ""

    if not llm_result or llm_result.get("status") != "ok":
        if llm_result:
            out["FirstPage_LLM_Status"] = as_text(llm_result.get("status"))
        return round(float(rule_score), 2), out

    llm_score = clamp_score(float(llm_result.get("relevance_score", 0.0)))
    llm_rule_pass = 1 if (
        int(llm_result.get("digital_twin_signal_count", 0)) >= 2
        or int(llm_result.get("answers_key_question_with_specifics", 0)) == 1
    ) else 0
    # Confidence-aware blending; when confidence is low, lean back to rule score.
    confidence = max(0.0, min(1.0, float(llm_result.get("confidence", 0.0))))
    adaptive_w = max(0.0, min(0.95, llm_weight * confidence))
    blended = round((1.0 - adaptive_w) * float(rule_score) + adaptive_w * llm_score, 2)

    out["FirstPage_Used_LLM"] = 1
    out["FirstPage_LLM_Status"] = "ok"
    out["FirstPage_Score_LLM"] = round(llm_score, 2)
    out["FirstPage_Score_Blend_Weight_LLM"] = round(adaptive_w, 3)
    out["FirstPage_LLM_DigitalTwinSignal_Count"] = int(llm_result.get("digital_twin_signal_count", 0))
    out["FirstPage_LLM_QuestionSpecifics_Rule"] = int(llm_result.get("answers_key_question_with_specifics", 0))
    out["FirstPage_LLM_Georgia_Anchor"] = int(llm_result.get("georgia_southeast_anchor", 0))
    out["FirstPage_LLM_EntitySpecificity_Count"] = int(llm_result.get("entity_specificity_count", 0))
    out["FirstPage_LLM_Recommendation"] = as_text(llm_result.get("recommendation"))
    out["FirstPage_LLM_Confidence"] = round(confidence, 3)
    out["FirstPage_LLM_KeySignals"] = ", ".join(llm_result.get("key_signals") or [])
    out["FirstPage_LLM_Rationale"] = as_text(llm_result.get("rationale"))
    out["FirstPage_LLM_Rule_Pass"] = llm_rule_pass
    out["FirstPage_LLM_Summary"] = as_text(llm_result.get("summary"))
    return blended, out


def stable_id_from_url(url: str) -> str:
    canonical = canonicalize_url(url)
    if not canonical:
        return "CAND_UNKNOWN"
    h = hashlib.sha1(canonical.encode("utf-8")).hexdigest()[:10]
    return f"CAND_{h.upper()}"


def probable_pdf(row: dict) -> bool:
    url = str(row.get("URL", "") or "")
    if is_pdf_url(url):
        return True
    file_type = as_text(row.get("File_Type")).lower()
    file_ext_guess = as_text(row.get("File_Ext_Guess")).lower()
    filename = as_text(row.get("Filename")).lower()
    if file_type == "pdf" or file_ext_guess == ".pdf" or filename.endswith(".pdf"):
        return True
    text = normalize_space(
        " ".join(
            [
                str(row.get("Title", "") or ""),
                str(row.get("Content_Snippet", "") or ""),
                str(row.get("Description", "") or ""),
                str(row.get("Query", "") or ""),
                filename_to_text(filename),
            ]
        )
    ).lower()
    return any(k in text for k in PDF_HINT_KEYWORDS)


def read_queries(path: Path, max_queries: int | None = None, query_offset: int = 0) -> list[str]:
    queries = [normalize_space(line) for line in path.read_text(encoding="utf-8").splitlines()]
    queries = [q for q in queries if q]
    if query_offset and query_offset > 0:
        queries = queries[query_offset:]
    if max_queries and max_queries > 0:
        return queries[:max_queries]
    return queries


def tavily_search_rows(
    *,
    queries: list[str],
    api_key: str,
    max_results: int,
    search_depth: str,
    golden_summary: str,
    query_mode: str = "hybrid",
    query_enhancement: bool = True,
    include_raw_content: bool = True,
) -> list[dict]:
    if TavilyClient is None:
        raise RuntimeError("tavily-python is not installed. Install tavily-python>=0.7.0 to run Tavily search.")
    client = TavilyClient(api_key=api_key)
    all_rows: list[dict] = []
    search_plan: list[tuple[int, str, str, str]] = []
    for q_idx, query in enumerate(queries, start=1):
        for variant_label, query_used in query_variants_for_mode(
            query,
            query_mode=query_mode,
            query_enhancement=query_enhancement,
        ):
            search_plan.append((q_idx, query, variant_label, query_used))

    for run_idx, (q_idx, query, variant_label, query_used) in enumerate(search_plan, start=1):
        print(f"[search {run_idx}/{len(search_plan)}] ({variant_label}) {query_used}")
        search_kwargs = {
            "query": query_used,
            "max_results": max_results,
            "search_depth": search_depth,
        }
        if include_raw_content:
            search_kwargs["include_raw_content"] = True
        try:
            response = client.search(**search_kwargs)
        except TypeError:
            # Backward compatibility for older tavily-python versions.
            search_kwargs.pop("include_raw_content", None)
            response = client.search(**search_kwargs)
        except Exception as exc:
            print(f"  search failed: {exc}")
            continue

        for rank, item in enumerate(response.get("results") or [], start=1):
            url_original = str(item.get("url") or "")
            url = canonicalize_url(url_original)
            title = str(item.get("title") or "")
            raw_content = normalize_space(str(item.get("raw_content") or ""))
            content_snippet = normalize_space(str(item.get("content") or ""))
            scoring_text = raw_content or content_snippet
            if not scoring_text:
                scoring_text = normalize_space(title)
            scoring_text = scoring_text[:2400]
            content_snippet = content_snippet[:1200]
            t_score_raw = float(item.get("score") or 0.0)
            domain = get_domain(url or url_original)
            file_ext = os.path.splitext(urlparse(url or url_original).path)[1].lower()

            sb = metadata_score_for_candidate(
                title=title,
                content=scoring_text,
                description=content_snippet,
                query=query_used,
                url=url or url_original,
                tavily_score_raw=t_score_raw,
                golden_summary=golden_summary,
            )

            row = {
                "Candidate_ID": stable_id_from_url(url or url_original),
                "Query_No": q_idx,
                "Query": query,
                "Query_Enhanced": query_used,
                "Query_Mode": variant_label,
                "Rank_In_Query": rank,
                "Title": title,
                "URL": url or url_original,
                "URL_Original": url_original,
                "URL_Canonical": url or url_original,
                "Source": domain,
                "Source_Domain": domain,
                "File_Ext_Guess": file_ext,
                "File_Type": "PDF" if file_ext == ".pdf" else "",
                "Is_PDF_URL": 1 if is_pdf_url(url or url_original) else 0,
                "Raw_Content_Available": 1 if raw_content else 0,
                "Raw_Content_Text": raw_content[:12000],
                "Content_Snippet": scoring_text,
                "Description": content_snippet or scoring_text,
                "Publication_Date_Metadata": "",
                "Tavily_Score_Raw": round(t_score_raw, 5),
                "Tavily_Score_Pct": round(max(0.0, min(100.0, t_score_raw * 100.0)), 2),
                "Metadata_Score": sb.metadata_score,
                "Signal_Category_Count": sb.category_hit_count,
                "Signal_Categories_Hit": sb.category_hits,
                "Signal_Keyword_Hits": sb.signal_keyword_hits,
                "Region_Hit": sb.region_hit,
                "ValueChain_Hits": sb.value_chain_hits,
                "PolicyLogistics_Hits": sb.policy_logistics_hits,
                "Question_Coverage": sb.question_coverage,
                "GeorgiaNode_Hits": sb.georgia_node_hits,
                "Specificity_Hits": sb.specificity_hits,
                "DocQuality_Hits": sb.document_quality_hits,
                "Metadata_TwoSignal_Rule": sb.two_signal_rule,
                "Metadata_QuestionSpecifics_Rule": sb.question_specifics_rule,
                "Source_Credibility": sb.source_credibility,
                "Penalty_Points": sb.penalty_points,
                "Lexical_Similarity": sb.lexical_similarity,
            }
            row["PDF_Candidate"] = 1 if probable_pdf(row) else 0
            row["Metadata_Rule_Pass"] = 1 if (
                int(row["Metadata_TwoSignal_Rule"]) == 1 or int(row["Metadata_QuestionSpecifics_Rule"]) == 1
            ) else 0
            all_rows.append(row)

    return all_rows


def metadata_rows_from_excel(
    *,
    input_xlsx: Path,
    golden_summary: str,
) -> list[dict]:
    df = pd.read_excel(input_xlsx)
    rows: list[dict] = []
    for idx, rec in enumerate(df.to_dict(orient="records"), start=1):
        url_original = as_text(rec.get("URL"))
        url = canonicalize_url(url_original)
        file_name = as_text(rec.get("Filename"))
        title = as_text(rec.get("Title")) or filename_to_text(file_name)
        title = title[:300]
        query = as_text(rec.get("Query"))
        if not query:
            query = normalize_space(
                " ".join(
                    [
                        as_text(rec.get("Category")),
                        as_text(rec.get("Industry")),
                        as_text(rec.get("Source")),
                    ]
                )
            )
        file_type = as_text(rec.get("File_Type"))
        file_text = filename_to_text(file_name)
        desc = as_text(rec.get("Description"))[:3000]
        if not desc:
            desc = normalize_space(
                " ".join(
                    [
                        title,
                        file_text,
                        file_name,
                        as_text(rec.get("Category")),
                        as_text(rec.get("Industry")),
                        as_text(rec.get("Source")),
                    ]
                )
            )[:3000]
        score_col = rec.get("Score")

        tavily_raw = 0.0
        try:
            s = float(score_col)
            tavily_raw = s / 100.0 if s > 1 else s
        except Exception:
            tavily_raw = 0.0

        sb = metadata_score_for_candidate(
            title=title,
            content=desc,
            description=desc,
            query=query,
            url=url or url_original,
            tavily_score_raw=tavily_raw,
            golden_summary=golden_summary,
        )

        candidate_id = stable_id_from_url(url or url_original or f"row-{idx}")
        file_ext = os.path.splitext(file_name)[1].lower() if file_name else os.path.splitext(urlparse(url or url_original).path)[1].lower()
        row = {
            "Candidate_ID": candidate_id,
            "Query_No": rec.get("Run_No", ""),
            "Query": query,
            "Query_Enhanced": query,
            "Query_Mode": as_text(rec.get("Query_Mode")) or "metadata_ingest",
            "Rank_In_Query": "",
            "Title": title,
            "URL": url or url_original,
            "URL_Original": url_original,
            "URL_Canonical": url or url_original,
            "Source": as_text(rec.get("Source")),
            "Source_Domain": get_domain(url or url_original),
            "File_Ext_Guess": file_ext,
            "File_Type": file_type,
            "Is_PDF_URL": 1 if is_pdf_url(url or url_original) else 0,
            "Content_Snippet": desc[:1200],
            "Description": desc,
            "Raw_Content_Text": "",
            "Filename": file_name,
            "Path": as_text(rec.get("Path")),
            "Existing_Path": as_text(rec.get("Path")),
            "Publication_Date_Metadata": as_text(rec.get("Date_Modified")) or as_text(rec.get("Date_Created")),
            "Tavily_Score_Raw": round(tavily_raw, 5),
            "Tavily_Score_Pct": round(max(0.0, min(100.0, tavily_raw * 100.0)), 2),
            "Metadata_Score": sb.metadata_score,
            "Signal_Category_Count": sb.category_hit_count,
            "Signal_Categories_Hit": sb.category_hits,
            "Signal_Keyword_Hits": sb.signal_keyword_hits,
            "Region_Hit": sb.region_hit,
            "ValueChain_Hits": sb.value_chain_hits,
            "PolicyLogistics_Hits": sb.policy_logistics_hits,
            "Question_Coverage": sb.question_coverage,
            "GeorgiaNode_Hits": sb.georgia_node_hits,
            "Specificity_Hits": sb.specificity_hits,
            "DocQuality_Hits": sb.document_quality_hits,
            "Metadata_TwoSignal_Rule": sb.two_signal_rule,
            "Metadata_QuestionSpecifics_Rule": sb.question_specifics_rule,
            "Source_Credibility": sb.source_credibility,
            "Penalty_Points": sb.penalty_points,
            "Lexical_Similarity": sb.lexical_similarity,
        }
        row["PDF_Candidate"] = 1 if probable_pdf(row) else 0
        row["Metadata_Rule_Pass"] = 1 if (
            int(row["Metadata_TwoSignal_Rule"]) == 1 or int(row["Metadata_QuestionSpecifics_Rule"]) == 1
        ) else 0
        rows.append(row)
    return rows


def dedupe_by_url_best_score(rows: list[dict]) -> list[dict]:
    best: dict[str, dict] = {}
    for row in rows:
        key = as_text(row.get("URL_Canonical")) or canonicalize_url(as_text(row.get("URL")))
        key = key.lower()
        if not key:
            filename_key = canonical_filename_key(as_text(row.get("Filename")))
            if filename_key:
                key = f"file::{filename_key}"
            else:
                title_key = normalize_space(as_text(row.get("Title"))).lower()
                key = f"title::{title_key}" if title_key else as_text(row.get("Candidate_ID"))
        prev = best.get(key)
        if not prev:
            best[key] = row
            continue
        if float(row.get("Metadata_Score", 0.0)) > float(prev.get("Metadata_Score", 0.0)):
            best[key] = row
    deduped = list(best.values())
    deduped.sort(key=lambda x: float(x.get("Metadata_Score", 0.0)), reverse=True)
    return deduped


def build_local_pdf_index(local_pdf_dirs: list[Path]) -> dict[str, str]:
    idx: dict[str, str] = {}
    for root in local_pdf_dirs:
        if not root.exists():
            continue
        for p in root.rglob("*.pdf"):
            idx[p.name.lower()] = str(p.resolve())
    return idx


def candidate_local_names(row: dict) -> list[str]:
    names: list[str] = []
    filename = str(row.get("Filename") or "").strip()
    if filename:
        names.append(filename)
    url = str(row.get("URL") or "")
    tail = os.path.basename(urlparse(url).path or "").strip()
    if tail:
        names.append(tail)
    title = str(row.get("Title") or "").strip()
    if title:
        names.append(f"{sanitize_filename(title)}.pdf")
    uniq: list[str] = []
    seen: set[str] = set()
    for n in names:
        k = n.lower()
        if k in seen:
            continue
        seen.add(k)
        uniq.append(n)
    return uniq


def parse_content_disposition(cd: str | None) -> str | None:
    if not cd:
        return None
    m = re.search(r"filename\*=UTF-8''([^;]+)", cd, flags=re.IGNORECASE)
    if m:
        return sanitize_filename(requests.utils.unquote(m.group(1)))
    m = re.search(r'filename="([^"]+)"', cd, flags=re.IGNORECASE)
    if m:
        return sanitize_filename(m.group(1))
    m = re.search(r"filename=([^;]+)", cd, flags=re.IGNORECASE)
    if m:
        return sanitize_filename(m.group(1).strip().strip('"'))
    return None


def download_pdf(
    *,
    url: str,
    output_dir: Path,
    fallback_name: str,
    timeout_sec: int,
    max_download_mb: float,
) -> tuple[str, float, str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "application/pdf,*/*;q=0.8",
    }
    tmp = output_dir / f"{fallback_name}.part"
    try:
        with requests.get(url, headers=headers, timeout=timeout_sec, stream=True, allow_redirects=True) as r:
            r.raise_for_status()
            content_type = (r.headers.get("Content-Type") or "").lower()
            final_url = r.url or url

            max_bytes = int(max_download_mb * 1024 * 1024)
            bytes_written = 0
            with tmp.open("wb") as f:
                for chunk in r.iter_content(chunk_size=1024 * 128):
                    if not chunk:
                        continue
                    f.write(chunk)
                    bytes_written += len(chunk)
                    if bytes_written > max_bytes:
                        try:
                            tmp.unlink(missing_ok=True)
                        except OSError:
                            pass
                        return "", 0.0, f"SkippedTooLarge>{max_download_mb}MB"

            size_mb = round(bytes_written / (1024 * 1024), 2)
            cd_name = parse_content_disposition(r.headers.get("Content-Disposition"))
            if cd_name:
                final_name = cd_name
            else:
                tail = os.path.basename(urlparse(final_url).path or "")
                final_name = sanitize_filename(tail) if tail else fallback_name
            if not final_name.lower().endswith(".pdf"):
                if "pdf" in content_type or is_pdf_url(final_url):
                    final_name = f"{os.path.splitext(final_name)[0]}.pdf"
                else:
                    final_name = f"{os.path.splitext(final_name)[0]}.pdf"

            final_path = output_dir / final_name
            if final_path.exists():
                b, e = os.path.splitext(final_name)
                k = 2
                while (output_dir / f"{b}_{k}{e}").exists():
                    k += 1
                final_path = output_dir / f"{b}_{k}{e}"

            os.replace(tmp, final_path)
            return str(final_path.resolve()), size_mb, "Downloaded"
    except Exception as exc:
        try:
            tmp.unlink(missing_ok=True)
        except OSError:
            pass
        return "", 0.0, f"DownloadFailed:{exc}"


def download_document(
    *,
    url: str,
    output_dir: Path,
    fallback_name: str,
    timeout_sec: int,
    max_download_mb: float,
) -> tuple[str, float, str, str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "*/*",
    }
    try:
        resp = requests.get(url, headers=headers, timeout=timeout_sec, allow_redirects=True)
        resp.raise_for_status()
        content_type = (resp.headers.get("Content-Type") or "").lower()
        body = resp.content
        if len(body) > int(max_download_mb * 1024 * 1024):
            return "", 0.0, f"SkippedTooLarge>{max_download_mb}MB", content_type

        cd_name = parse_content_disposition(resp.headers.get("Content-Disposition"))
        tail = os.path.basename(urlparse(resp.url or url).path or "")
        final_name = cd_name or sanitize_filename(tail or fallback_name or "candidate")
        ext = os.path.splitext(final_name)[1].lower()
        if "pdf" in content_type and ext != ".pdf":
            ext = ".pdf"
        elif "html" in content_type and ext not in {".html", ".htm"}:
            ext = ".html"
        elif "text/plain" in content_type and ext != ".txt":
            ext = ".txt"
        elif not ext:
            ext = ".pdf" if is_pdf_url(url) else ".html"
        final_name = f"{os.path.splitext(final_name)[0]}{ext}"

        final_path = output_dir / final_name
        if final_path.exists():
            stem, suffix = os.path.splitext(final_name)
            k = 2
            while (output_dir / f"{stem}_{k}{suffix}").exists():
                k += 1
            final_path = output_dir / f"{stem}_{k}{suffix}"
        final_path.write_bytes(body)
        size_mb = round(len(body) / (1024 * 1024), 2)
        return str(final_path.resolve()), size_mb, "Downloaded", content_type
    except Exception as exc:
        return "", 0.0, f"DownloadFailed:{exc}", ""


def write_jsonl(path: Path, rows: list[dict]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            safe_row = {k: v for k, v in row.items() if not str(k).startswith("_")}
            handle.write(json.dumps(safe_row, ensure_ascii=True) + "\n")


def maybe_write_excel(enabled: bool, df: pd.DataFrame, path: Path) -> None:
    if enabled:
        write_excel(df, path)


def maybe_write_jsonl(enabled: bool, path: Path, rows: list[dict]) -> None:
    if enabled:
        write_jsonl(path, rows)


def llm_document_judge(
    *,
    cfg: LLMConfig,
    golden_summary: str,
    card_row: dict[str, Any],
) -> dict[str, Any]:
    system_prompt = (
        "You are an expert evaluator for GNEM phase-1 RAG corpus construction. "
        "Judge whether the document should be kept for human-reviewed supply-chain RAG ingestion. "
        "Return strict JSON only."
    )
    user_prompt = textwrap.dedent(
        f"""
        GNEM_USE_CASE:
        {golden_summary}

        DOCUMENT_CARD:
        - card_level: {as_text(card_row.get("Card_Level"))}
        - title: {as_text(card_row.get("Title"))}
        - source_domain: {as_text(card_row.get("Source_Domain"))}
        - file_type: {as_text(card_row.get("File_Type"))}
        - publication_date: {as_text(card_row.get("Publication_Date"))}
        - source: {as_text(card_row.get("Document_Source"))}
        - metadata_summary: {as_text(card_row.get("Metadata_Summary"))}
        - first_page_summary: {as_text(card_row.get("FirstPage_Summary"))}
        - first_two_pages_summary: {as_text(card_row.get("FirstTwoPages_Summary"))}
        - sampled_page_summaries: {as_text(card_row.get("Sampled_Page_Summaries"))}
        - headings: {as_text(card_row.get("Headings_TOC"))}
        - top_evidence_snippets: {as_text(card_row.get("Top_Evidence_Snippets"))}
        - evidence_page_numbers: {as_text(card_row.get("Evidence_Page_Numbers"))}
        - extracted_companies: {as_text(card_row.get("Extracted_Companies"))}
        - extracted_counties: {as_text(card_row.get("Extracted_Counties"))}
        - extracted_oems: {as_text(card_row.get("Extracted_OEMs"))}
        - extracted_ports: {as_text(card_row.get("Extracted_Ports"))}
        - extracted_facilities: {as_text(card_row.get("Extracted_Facilities"))}
        - extracted_capacities: {as_text(card_row.get("Extracted_Capacities"))}
        - extracted_dates: {as_text(card_row.get("Extracted_Dates"))}
        - heuristic_score: {as_text(card_row.get("Heuristic_Score"))}
        - semantic_embedding_score: {as_text(card_row.get("Semantic_Embedding_Score", card_row.get("Embedding_Score")))}
        - lexical_entity_score: {as_text(card_row.get("Lexical_Entity_Score"))}
        - hybrid_score: {as_text(card_row.get("Hybrid_Score"))}
        - direct_usecase_score: {as_text(card_row.get("Direct_Usecase_Score"))}
        - adjacent_background_score: {as_text(card_row.get("Adjacent_Background_Score"))}
        - research_only_score: {as_text(card_row.get("Research_Only_Score"))}
        - generic_news_score: {as_text(card_row.get("Generic_News_Score"))}
        - marketing_noise_score: {as_text(card_row.get("Marketing_Noise_Score"))}

        Return a JSON object with exactly these keys:
        {{
          "relevance_score": 0-10,
          "usecase_match": 0-10,
          "information_quality": 0-10,
          "noise_level": 0-10,
          "decision": "keep" or "review" or "discard",
          "reason": "max 80 words",
          "top_evidence_used": ["...", "...", "..."],
          "confidence": 0-1
        }}
        """
    ).strip()

    data, status = llm_chat_json(cfg=cfg, system_prompt=system_prompt, user_prompt=user_prompt)
    if not data:
        return {"LLM_Judge_Status": status}

    top_evidence_used = data.get("top_evidence_used")
    if not isinstance(top_evidence_used, list):
        top_evidence_used = []
    try:
        confidence = float(data.get("confidence", 0.0))
    except Exception:
        confidence = 0.0

    relevance = max(0.0, min(10.0, float(data.get("relevance_score", 0.0))))
    usecase_match = max(0.0, min(10.0, float(data.get("usecase_match", 0.0))))
    information_quality = max(0.0, min(10.0, float(data.get("information_quality", 0.0))))
    noise_level = max(0.0, min(10.0, float(data.get("noise_level", 0.0))))
    judge_pass = int(relevance >= 8 and usecase_match >= 8 and information_quality >= 7 and noise_level <= 3)
    model_decision = as_text(data.get("decision")).lower()
    if judge_pass:
        derived_decision = "keep"
    elif relevance >= 7 and usecase_match >= 6 and information_quality >= 6 and noise_level <= 5:
        derived_decision = "review"
    else:
        derived_decision = "discard"
    return {
        "LLM_Judge_Status": "ok",
        "LLM_Relevance_Score": round(relevance, 2),
        "LLM_Usecase_Match": round(usecase_match, 2),
        "LLM_Information_Quality": round(information_quality, 2),
        "LLM_Noise_Level": round(noise_level, 2),
        "LLM_Model_Decision": model_decision,
        "LLM_Decision": derived_decision,
        "LLM_Reason": as_text(data.get("reason"))[:500],
        "LLM_Top_Evidence_Used": " || ".join(as_text(x) for x in top_evidence_used if as_text(x))[:1500],
        "LLM_Confidence": round(max(0.0, min(1.0, confidence)), 3),
        "LLM_Judge_Pass": judge_pass,
    }


def extract_first_page_text(pdf_path: str) -> str:
    if PdfReader:
        reader = PdfReader(pdf_path)
        if not reader.pages:
            return ""
        txt = reader.pages[0].extract_text() or ""
        return normalize_space(txt)
    if fitz:
        doc = fitz.open(pdf_path)
        try:
            if doc.page_count == 0:
                return ""
            txt = doc.load_page(0).get_text("text") or ""
            return normalize_space(txt)
        finally:
            doc.close()
    raise RuntimeError("No PDF text extractor available. Install pypdf>=5.0.0 or PyMuPDF.")


def ensure_columns(df: pd.DataFrame, preferred: list[str]) -> pd.DataFrame:
    working = df.copy()
    for column in preferred:
        if column not in working.columns:
            working[column] = ""
    return working[preferred]


def sanitize_excel_value(value: Any) -> Any:
    if isinstance(value, str):
        return ILLEGAL_XLSX_CHARS_RE.sub("", value)[:32767]
    return value


def sanitize_dataframe_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    if hasattr(df, "map"):
        return df.map(sanitize_excel_value)
    return df.apply(lambda col: col.map(sanitize_excel_value))


def format_excel_workbook(path: Path) -> None:
    if load_workbook is None or Font is None or PatternFill is None or Alignment is None or get_column_letter is None:
        return
    try:
        workbook = load_workbook(path)
    except Exception:
        return
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_idx in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(column_idx)
            max_length = 0
            for cell in worksheet[column_letter]:
                try:
                    cell_length = len(str(cell.value or ""))
                except Exception:
                    cell_length = 0
                if cell_length > max_length:
                    max_length = cell_length
            worksheet.column_dimensions[column_letter].width = min(max(12, max_length + 2), 48)
    workbook.save(path)


def write_excel(df: pd.DataFrame, path: Path) -> None:
    safe_df = sanitize_dataframe_for_excel(df)
    safe_df.to_excel(path, index=False)
    format_excel_workbook(path)


def write_excel_sheets(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            sanitize_dataframe_for_excel(df).to_excel(writer, sheet_name=sheet_name[:31], index=False)
    format_excel_workbook(path)


def resolve_existing_repo_path(primary: str | Path, *fallbacks: str | Path) -> Path | None:
    primary_path = Path(primary)
    if primary_path.exists():
        return primary_path
    names = [str(candidate) for candidate in fallbacks if str(candidate)]
    if not names:
        return None
    return detect_grounding_file(names, Path.cwd())


def safe_float(value: object, default: float = 0.0) -> float:
    try:
        return float(value)
    except Exception:
        return default


FINAL_EXPORT_COLUMNS = [
    "Candidate_ID",
    "Final_Decision",
    "Title",
    "Source",
    "File_Path_Or_URL",
    "Curated_File_Path",
    "Query",
    "Query_Mode",
    "Primary_Subtopic",
    "Publication_Date",
    "File_Type",
    "Final_Rank_Score",
    "Metadata_Score",
    "Heuristic_Score",
    "Hybrid_Score",
    "Rerank_Score",
    "Direct_Usecase_Score",
    "Adjacent_Background_Score",
    "Credibility_Score",
    "LLM_Decision",
    "LLM_Relevance_Score",
    "LLM_Usecase_Match",
    "LLM_Information_Quality",
    "Evidence_Page_Numbers",
    "Top_Evidence_Snippets",
    "Keep_Review_Discard_Reason",
    "Extracted_Companies",
    "Extracted_Counties",
    "Extracted_OEMs",
    "Extracted_Ports",
    "Extracted_Facilities",
    "Extracted_Capacities",
    "Extracted_Dates",
    "Curated_Copy_Status",
]

REJECTED_EXPORT_COLUMNS = [
    "Candidate_ID",
    "Title",
    "Source",
    "File_Path_Or_URL",
    "Query",
    "Query_Mode",
    "Metadata_Score",
    "Heuristic_Score",
    "Hybrid_Score",
    "Rerank_Score",
    "Direct_Usecase_Score",
    "Adjacent_Background_Score",
    "Credibility_Score",
    "Stage2_Filter_Pass",
    "Source_File_Available",
    "Shortlist_Pass",
    "Exact_Duplicate_Flag",
    "Duplicate_Of_Candidate_ID",
    "Near_Duplicate_Flag",
    "Near_Duplicate_Of_Candidate_ID",
    "Source_Blocklist_Flag",
    "Source_LowTrust_Flag",
    "Acquisition_Status",
    "Keep_Review_Discard_Reason",
    "Final_Decision",
]

DEFAULT_EV_AUTOMATION_READY_DIR = (
    Path(__file__).resolve().parent.parent
    / "evAutomationUpdated"
    / "data"
    / "tavily ready documents"
)


def existing_file_path(*values: object) -> str:
    for value in values:
        path_str = as_text(value)
        if not path_str:
            continue
        try:
            path = Path(path_str)
        except Exception:
            continue
        if path.exists() and path.is_file():
            return str(path.resolve())
    return ""


def repo_relative_posix(path: Path, root: Path) -> str:
    try:
        return path.resolve().relative_to(root.resolve()).as_posix()
    except Exception:
        return path.name


def populate_source_file_fields(row: dict[str, Any]) -> dict[str, Any]:
    source_file_path = existing_file_path(
        row.get("Curated_File_Path"),
        row.get("Document_File_Path"),
        row.get("Acquired_File_Path"),
        row.get("Resolved_File_Path"),
    )
    row["Source_File_Path"] = source_file_path
    row["Source_File_Available"] = 1 if source_file_path else 0
    return row


def stage2_rejection_reason(row: dict[str, Any]) -> str:
    if int(row.get("Document_Candidate", 0)) != 1:
        return "Rejected before detailed scoring because the candidate did not look like a usable document."
    if int(row.get("Metadata_Pass", 0)) != 1:
        return (
            f"Rejected at Stage 2 because Metadata_Score={round(safe_float(row.get('Metadata_Score')), 2)} "
            f"did not clear the threshold."
        )
    if int(row.get("Metadata_Rule_Pass", 0)) != 1:
        return "Rejected at Stage 2 because the metadata rule gate did not pass."
    if int(row.get("Source_File_Available", 0)) != 1:
        return "Rejected because the source file could not be acquired or preserved locally."
    return "Rejected before detailed scoring due to Stage 2 filtering."


def eligible_for_review_promotion(
    row: dict[str, Any],
    *,
    credibility_threshold: float,
    heuristic_threshold: float,
    hybrid_threshold: float,
    direct_usecase_threshold: float,
) -> bool:
    if as_text(row.get("Final_Decision")) != "discard":
        return False
    if int(row.get("Stage2_Filter_Pass", 0)) != 1 or int(row.get("Source_File_Available", 0)) != 1:
        return False
    if int(row.get("Duplicate_Master_Flag", 1)) != 1 or int(row.get("Near_Duplicate_Master_Flag", 1)) != 1:
        return False
    if int(row.get("Source_Blocklist_Flag", 0)) == 1:
        return False
    if int(row.get("Source_LowTrust_Flag", 0)) == 1 and safe_float(row.get("Credibility_Score")) < 75.0:
        return False
    if safe_float(row.get("Credibility_Score")) < max(45.0, credibility_threshold - 10.0):
        return False
    if safe_float(row.get("Heuristic_Score")) < max(40.0, heuristic_threshold - 5.0):
        return False
    if safe_float(row.get("Rerank_Score")) < 55.0:
        return False
    if safe_float(row.get("Direct_Usecase_Score")) >= max(0.45, direct_usecase_threshold - 0.15):
        return True
    if (
        safe_float(row.get("Adjacent_Background_Score")) >= 0.60
        and safe_float(row.get("Hybrid_Score")) >= max(45.0, hybrid_threshold - 15.0)
    ):
        return True
    return (
        safe_float(row.get("Hybrid_Score")) >= max(50.0, hybrid_threshold - 15.0)
        and safe_float(row.get("Embedding_Rerank_Score")) >= 45.0
    )


def final_rank_sort_key(row: dict[str, Any]) -> tuple[float, float, float, float]:
    return (
        safe_float(row.get("Final_Rank_Score")),
        safe_float(row.get("Rerank_Score")),
        safe_float(row.get("Hybrid_Score")),
        safe_float(row.get("Credibility_Score")),
    )


def publish_curated_documents_to_ready_dir(
    rows: list[dict[str, Any]],
    ready_dir: Path,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    ready_dir.mkdir(parents=True, exist_ok=True)
    keep_dir = ready_dir / "keep"
    review_dir = ready_dir / "review"
    keep_dir.mkdir(parents=True, exist_ok=True)
    review_dir.mkdir(parents=True, exist_ok=True)

    manifest_columns = [
        "Row_Number",
        "Candidate_ID",
        "Title",
        "Final_Decision",
        "URL",
        "FilePath",
        "Filename",
        "Retrieved_At",
        "Match_Status",
    ]
    unmatched_columns = [
        "Row_Number",
        "Candidate_ID",
        "Title",
        "Final_Decision",
        "URL",
        "Match_Status",
    ]
    manifest_rows: list[dict[str, Any]] = []
    unmatched_rows: list[dict[str, Any]] = []
    repo_root = ready_dir.parent.parent.parent if ready_dir.parent.name == "data" else ready_dir
    retrieved_at = datetime.now().isoformat(timespec="seconds")

    for row_number, row in enumerate(rows, start=2):
        decision = as_text(row.get("Final_Decision"))
        if decision not in {"keep", "review"}:
            continue

        source_path = as_text(row.get("Curated_File_Path")) or as_text(row.get("Document_File_Path"))
        if not source_path or not Path(source_path).exists():
            unmatched_rows.append(
                {
                    "Row_Number": row_number,
                    "Candidate_ID": as_text(row.get("Candidate_ID")),
                    "Title": as_text(row.get("Title")),
                    "Final_Decision": decision,
                    "URL": as_text(row.get("Document_URL")) or as_text(row.get("URL")),
                    "Match_Status": "missing_source",
                }
            )
            continue

        source = Path(source_path)
        target_dir = keep_dir if decision == "keep" else review_dir
        target_name = safe_output_filename(target_dir, source.name)
        target = target_dir / target_name
        suffix = source.suffix
        stem = target.stem
        attempt = 2
        while target.exists() and target.resolve() != source.resolve():
            target = target_dir / safe_output_filename(target_dir, f"{stem}_{attempt}{suffix}")
            attempt += 1
        if not target.exists():
            shutil.copy2(source, target)

        manifest_rows.append(
            {
                "Row_Number": row_number,
                "Candidate_ID": as_text(row.get("Candidate_ID")),
                "Title": as_text(row.get("Title")),
                "Final_Decision": decision,
                "URL": as_text(row.get("Document_URL")) or as_text(row.get("URL")),
                "FilePath": repo_relative_posix(target, repo_root),
                "Filename": target.name,
                "Retrieved_At": retrieved_at,
                "Match_Status": "published",
            }
        )

    manifest_path = ready_dir / "tavily_ready_documents_manifest.csv"
    unmatched_path = ready_dir / "tavily_ready_documents_unmatched.csv"
    pd.DataFrame(manifest_rows, columns=manifest_columns).to_csv(manifest_path, index=False)
    pd.DataFrame(unmatched_rows, columns=unmatched_columns).to_csv(unmatched_path, index=False)
    return manifest_rows, unmatched_rows


def compute_file_sha256(path_str: str) -> str:
    if not path_str:
        return ""
    path = Path(path_str)
    if not path.exists() or not path.is_file():
        return ""
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def compute_text_sha256(text: str) -> str:
    normalized = normalize_space(text)
    if not normalized:
        return ""
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()


def publication_date_rank(value: object) -> float:
    text = as_text(value)
    if not text:
        return 0.0
    try:
        parsed = pd.to_datetime(text, errors="coerce", utc=True)
    except Exception:
        return 0.0
    if pd.isna(parsed):
        return 0.0
    try:
        return float(parsed.timestamp())
    except Exception:
        return 0.0


def duplicate_rank_key(row: dict[str, Any]) -> tuple[float, float, float, float, float, float]:
    return (
        publication_date_rank(row.get("Publication_Date")),
        safe_float(row.get("Rerank_Score")),
        safe_float(row.get("Hybrid_Score")),
        safe_float(row.get("Heuristic_Score")),
        safe_float(row.get("Metadata_Score")),
        safe_float(row.get("Document_Content_Chars")),
    )


def normalized_duplicate_text(row: dict[str, Any]) -> str:
    text = normalize_space(
        as_text(row.get("_full_text"))
        or as_text(row.get("_card_text"))
        or " ".join(
            [
                as_text(row.get("Title")),
                as_text(row.get("Description")),
                as_text(row.get("Content_Snippet")),
            ]
        )
    )
    if "<" in text and ">" in text:
        text = re.sub(r"(?is)<(script|style|noscript).*?>.*?</\1>", " ", text)
        text = re.sub(r"(?is)<[^>]+>", " ", text)
        text = normalize_space(text)
    return re.sub(r"\s+", " ", text.lower())[:90000]


def simhash_signature(text: str, bits: int = 64) -> int:
    tokens = re.findall(r"[a-z0-9][a-z0-9\-\+\.]*", text.lower())
    if not tokens:
        return 0
    grams = list(tokens)
    grams.extend(f"{tokens[idx]}__{tokens[idx + 1]}" for idx in range(len(tokens) - 1))
    weights = [0] * bits
    for gram in grams:
        digest = hashlib.sha1(gram.encode("utf-8")).hexdigest()
        value = int(digest[:16], 16)
        for bit_idx in range(bits):
            weights[bit_idx] += 1 if value & (1 << bit_idx) else -1
    signature = 0
    for bit_idx, weight in enumerate(weights):
        if weight >= 0:
            signature |= 1 << bit_idx
    return signature


def hamming_distance(left: int, right: int) -> int:
    return (left ^ right).bit_count()


def summarize_final_supporting_factors(row: dict[str, Any]) -> tuple[str, str, str]:
    factors: list[str] = []
    source_domain = as_text(row.get("Source_Domain"))
    final_decision = as_text(row.get("Final_Decision")).lower()
    hybrid = round(safe_float(row.get("Hybrid_Score")), 2)
    credibility = round(safe_float(row.get("Credibility_Score")), 2)
    direct = round(safe_float(row.get("Direct_Usecase_Score")) * 100.0, 1)
    metadata = round(safe_float(row.get("Metadata_Score")), 2)
    primary_subtopic = as_text(row.get("Primary_Subtopic"))

    if source_domain:
        factors.append(f"source={source_domain}")
    if primary_subtopic:
        factors.append(f"subtopic={primary_subtopic}")
    if hybrid > 0:
        factors.append(f"hybrid={hybrid}")
    if credibility > 0:
        factors.append(f"credibility={credibility}")
    if direct > 0:
        factors.append(f"direct_usecase_pct={direct}")
    if metadata > 0:
        factors.append(f"metadata={metadata}")
    if as_text(row.get("Evidence_Page_Numbers")):
        factors.append(f"evidence_pages={as_text(row.get('Evidence_Page_Numbers'))}")
    if as_text(row.get("Extracted_Companies")):
        factors.append(f"companies={as_text(row.get('Extracted_Companies'))[:120]}")
    if as_text(row.get("Extracted_Counties")):
        factors.append(f"counties={as_text(row.get('Extracted_Counties'))[:120]}")
    if as_text(row.get("Extracted_OEMs")):
        factors.append(f"oems={as_text(row.get('Extracted_OEMs'))[:120]}")
    if as_text(row.get("Extracted_Capacities")):
        factors.append(f"capacities={as_text(row.get('Extracted_Capacities'))[:120]}")
    if int(row.get("Near_Duplicate_Flag", 0)) == 1 and int(row.get("Near_Duplicate_Master_Flag", 1)) == 1:
        factors.append("selected_as_best_near_duplicate_version")

    support_summary = "; ".join(factors)[:900]
    decision_reason = as_text(row.get("Keep_Review_Discard_Reason"))[:700]

    rationale_parts: list[str] = []
    if final_decision in {"keep", "review"}:
        rationale_parts.append(f"Final decision={final_decision}")
        if hybrid >= 65.0:
            rationale_parts.append("strong hybrid relevance")
        elif hybrid >= 55.0:
            rationale_parts.append("moderate hybrid relevance")
        if credibility >= 75.0:
            rationale_parts.append("high credibility")
        elif credibility >= 60.0:
            rationale_parts.append("acceptable credibility")
        if direct >= 70.0:
            rationale_parts.append("direct GNEM use-case fit")
        elif direct >= 55.0:
            rationale_parts.append("borderline direct-use-case fit")
        if as_text(row.get("Evidence_Page_Numbers")):
            rationale_parts.append("page-linked evidence available")
        if as_text(row.get("Extracted_Companies")) or as_text(row.get("Extracted_Counties")) or as_text(row.get("Extracted_OEMs")):
            rationale_parts.append("grounded Georgia-specific entities present")
    else:
        rationale_parts.append(f"Final decision={final_decision}")
        if decision_reason:
            rationale_parts.append(decision_reason)

    final_rationale = ". ".join(part for part in rationale_parts if part)[:700]
    return final_rationale, support_summary, decision_reason


def apply_exact_duplicate_pass(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        canonical_url = as_text(row.get("URL_Canonical")) or canonicalize_url(as_text(row.get("URL")))
        file_path = (
            as_text(row.get("Document_File_Path"))
            or as_text(row.get("Acquired_File_Path"))
            or as_text(row.get("Resolved_File_Path"))
        )
        text_hash = compute_text_sha256(as_text(row.get("_full_text")))
        bytes_hash = compute_file_sha256(file_path)
        row["URL_Canonical"] = canonical_url
        row["Document_Text_SHA256"] = text_hash
        row["Document_Bytes_SHA256"] = bytes_hash
        row["Exact_Duplicate_Flag"] = 0
        row["Duplicate_Master_Flag"] = 1
        row["Duplicate_Of_Candidate_ID"] = ""
        row["Duplicate_Group_Key"] = ""
        row["Duplicate_Group_Type"] = ""
        dedupe_key = ""
        dedupe_type = ""
        if bytes_hash:
            dedupe_key = f"bytes::{bytes_hash}"
            dedupe_type = "bytes_sha256"
        elif text_hash:
            dedupe_key = f"text::{text_hash}"
            dedupe_type = "text_sha256"
        elif canonical_url:
            dedupe_key = f"url::{canonical_url.lower()}"
            dedupe_type = "url_canonical"
        if dedupe_key:
            row["Duplicate_Group_Key"] = dedupe_key
            row["Duplicate_Group_Type"] = dedupe_type
            groups.setdefault(dedupe_key, []).append(row)

    for group_key, members in groups.items():
        if len(members) <= 1:
            continue
        master = max(members, key=duplicate_rank_key)
        for row in members:
            row["Exact_Duplicate_Flag"] = 1
            row["Duplicate_Master_Flag"] = 1 if row is master else 0
            row["Duplicate_Of_Candidate_ID"] = "" if row is master else as_text(master.get("Candidate_ID"))
            row["Duplicate_Group_Key"] = group_key
            if row is not master:
                row["Preliminary_Shortlist_Pass"] = 0
                row["Shortlist_Pass"] = 0
                row["Shortlist_Selected_By_Ratio"] = 0
    return rows


def apply_near_duplicate_pass(
    rows: list[dict[str, Any]],
    *,
    max_hamming_distance: int = 4,
    min_text_similarity: float = 0.88,
    min_title_similarity: float = 0.60,
) -> list[dict[str, Any]]:
    for row in rows:
        row["Near_Duplicate_Flag"] = 0
        row["Near_Duplicate_Master_Flag"] = 1
        row["Near_Duplicate_Of_Candidate_ID"] = ""
        row["Near_Duplicate_Group_Key"] = ""
        row["Near_Duplicate_Min_Hamming"] = ""
        row["Near_Duplicate_Text_Similarity"] = ""
        row["Near_Duplicate_Title_Similarity"] = ""

    candidates: list[dict[str, Any]] = []
    for row in rows:
        if int(row.get("Exact_Duplicate_Flag", 0)) == 1 and int(row.get("Duplicate_Master_Flag", 1)) == 0:
            continue
        duplicate_text = normalized_duplicate_text(row)
        if len(duplicate_text) < 200:
            continue
        title_text = normalize_space(
            as_text(row.get("Title"))
            or filename_to_text(as_text(row.get("Filename")))
            or Path(as_text(row.get("Document_File_Path")) or "document").stem
        )
        row["_near_dup_text"] = duplicate_text
        row["_near_dup_title"] = title_text
        row["_near_dup_signature"] = simhash_signature(duplicate_text)
        candidates.append(row)

    if len(candidates) <= 1:
        return rows

    parents = list(range(len(candidates)))

    def find(index: int) -> int:
        while parents[index] != index:
            parents[index] = parents[parents[index]]
            index = parents[index]
        return index

    def union(left_index: int, right_index: int) -> None:
        left_root = find(left_index)
        right_root = find(right_index)
        if left_root != right_root:
            parents[right_root] = left_root

    for left_index, left_row in enumerate(candidates):
        left_text = as_text(left_row.get("_near_dup_text"))
        left_title = as_text(left_row.get("_near_dup_title"))
        left_signature = int(left_row.get("_near_dup_signature") or 0)
        for right_index in range(left_index + 1, len(candidates)):
            right_row = candidates[right_index]
            right_text = as_text(right_row.get("_near_dup_text"))
            length_ratio = min(len(left_text), len(right_text)) / max(1, max(len(left_text), len(right_text)))
            if length_ratio < 0.65:
                continue
            title_similarity = cosine_similarity_text(left_title, as_text(right_row.get("_near_dup_title")))
            distance = hamming_distance(left_signature, int(right_row.get("_near_dup_signature") or 0))
            if distance > max_hamming_distance and title_similarity < min_title_similarity:
                continue
            text_similarity = cosine_similarity_text(left_text[:20000], right_text[:20000])
            strong_match = text_similarity >= min_text_similarity
            paired_match = distance <= 2 and title_similarity >= min_title_similarity
            blended_match = text_similarity >= 0.84 and title_similarity >= 0.80
            extended_hash_match = distance <= 12 and text_similarity >= 0.82
            if strong_match or paired_match or blended_match or extended_hash_match:
                union(left_index, right_index)

    grouped: dict[int, list[dict[str, Any]]] = {}
    for index, row in enumerate(candidates):
        grouped.setdefault(find(index), []).append(row)

    for members in grouped.values():
        if len(members) <= 1:
            continue
        master = max(members, key=duplicate_rank_key)
        group_key = f"near::{as_text(master.get('Candidate_ID')) or hashlib.sha1(as_text(master.get('Title')).encode('utf-8')).hexdigest()[:12]}"
        master_text = as_text(master.get("_near_dup_text"))
        master_title = as_text(master.get("_near_dup_title"))
        master_signature = int(master.get("_near_dup_signature") or 0)
        for row in members:
            distance = hamming_distance(master_signature, int(row.get("_near_dup_signature") or 0))
            title_similarity = cosine_similarity_text(master_title, as_text(row.get("_near_dup_title")))
            text_similarity = cosine_similarity_text(master_text[:20000], as_text(row.get("_near_dup_text"))[:20000])
            row["Near_Duplicate_Flag"] = 1
            row["Near_Duplicate_Master_Flag"] = 1 if row is master else 0
            row["Near_Duplicate_Of_Candidate_ID"] = "" if row is master else as_text(master.get("Candidate_ID"))
            row["Near_Duplicate_Group_Key"] = group_key
            row["Near_Duplicate_Min_Hamming"] = distance
            row["Near_Duplicate_Text_Similarity"] = round(text_similarity, 4)
            row["Near_Duplicate_Title_Similarity"] = round(title_similarity, 4)
            if row is not master:
                row["Preliminary_Shortlist_Pass"] = 0
                row["Shortlist_Pass"] = 0
                row["Shortlist_Selected_By_Ratio"] = 0

    for row in candidates:
        row.pop("_near_dup_text", None)
        row.pop("_near_dup_title", None)
        row.pop("_near_dup_signature", None)
    return rows


def subtopic_category_mapping(primary_subtopic: str) -> tuple[str, str]:
    mapping = {
        "who_where_scale": ("Supply Chain Evidence", "Facilities and Capacity"),
        "supplier_tiering": ("Supply Chain Evidence", "Supplier and Tiering"),
        "logistics_risk": ("Infrastructure and Risk", "Logistics and Risk"),
        "policy_localization": ("Policy and Localization", "Incentives and Localization"),
        "recycling_materials": ("Supply Chain Evidence", "Recycling and Materials"),
    }
    return mapping.get(primary_subtopic, ("Supply Chain Evidence", "General"))


def document_registry_dataframe(rows: list[dict[str, Any]]) -> pd.DataFrame:
    today = datetime.now().strftime("%Y-%m-%d")
    columns = [
        "Document_ID",
        "Document_Name",
        "Document_Type",
        "Category",
        "Sub_Category",
        "Source",
        "Source_URL",
        "File_Path",
        "File_Size_MB",
        "Page_Count",
        "Date_Published",
        "Date_Acquired",
        "Date_Processed",
        "Processing_Status",
        "Language",
        "Contains_Tables",
        "Contains_Images",
        "Confidentiality",
        "Retention_Period",
        "Owner",
        "Quality_Score",
        "Final_Decision",
        "Final_Selection_Rationale",
        "Supporting_Factors",
        "Notes",
    ]
    registry_rows: list[dict[str, Any]] = []
    for row in rows:
        primary_subtopic = as_text(row.get("Primary_Subtopic"))
        category, sub_category = subtopic_category_mapping(primary_subtopic)
        final_rationale, support_summary, decision_reason = summarize_final_supporting_factors(row)
        file_path = (
            as_text(row.get("Curated_File_Path"))
            or as_text(row.get("Document_File_Path"))
            or as_text(row.get("Acquired_File_Path"))
        )
        document_name = Path(file_path).name if file_path else sanitize_filename(as_text(row.get("Title")) or as_text(row.get("Candidate_ID")))
        file_size_mb = safe_float(row.get("Acquired_Size_MB"))
        if not file_size_mb and file_path and Path(file_path).exists():
            try:
                file_size_mb = round(Path(file_path).stat().st_size / (1024 * 1024), 2)
            except OSError:
                file_size_mb = 0.0
        registry_rows.append(
            {
                "Document_ID": as_text(row.get("Candidate_ID")),
                "Document_Name": document_name,
                "Document_Type": as_text(row.get("File_Type")) or "Unknown",
                "Category": category,
                "Sub_Category": sub_category,
                "Source": as_text(row.get("Source_Domain")) or as_text(row.get("Source")) or "Unknown",
                "Source_URL": as_text(row.get("Document_URL")) or as_text(row.get("URL")),
                "File_Path": file_path,
                "File_Size_MB": round(file_size_mb, 2) if file_size_mb else "",
                "Page_Count": int(safe_float(row.get("Document_Total_Pages"))) if safe_float(row.get("Document_Total_Pages")) > 0 else "",
                "Date_Published": as_text(row.get("Publication_Date")),
                "Date_Acquired": today if file_path else "",
                "Date_Processed": today,
                "Processing_Status": f"Curated-{as_text(row.get('Final_Decision')).title()}",
                "Language": as_text(row.get("Language")) or "Unknown",
                "Contains_Tables": "Unknown",
                "Contains_Images": "Unknown",
                "Confidentiality": "Public",
                "Retention_Period": "TBD",
                "Owner": "GNEM Research Pipeline",
                "Quality_Score": round(safe_float(row.get("Final_Rank_Score")), 2),
                "Final_Decision": as_text(row.get("Final_Decision")).title(),
                "Final_Selection_Rationale": final_rationale,
                "Supporting_Factors": support_summary,
                "Notes": normalize_space(
                    " ".join(
                        [
                            decision_reason,
                            f"Credibility={as_text(row.get('Credibility_Score'))}" if as_text(row.get("Credibility_Score")) else "",
                            f"Hybrid={as_text(row.get('Hybrid_Score'))}" if as_text(row.get("Hybrid_Score")) else "",
                            f"LLM={as_text(row.get('LLM_Decision'))}" if as_text(row.get("LLM_Decision")) else "",
                            f"CopyStatus={as_text(row.get('Curated_Copy_Status'))}" if as_text(row.get("Curated_Copy_Status")) else "",
                        ]
                    )
                )[:500],
            }
        )
    return pd.DataFrame(registry_rows, columns=columns)


def chunk_registry_dataframe(rows: list[dict[str, Any]]) -> pd.DataFrame:
    today = datetime.now().strftime("%Y-%m-%d")
    columns = [
        "Chunk_ID",
        "Document_ID",
        "Category",
        "Entity_Type",
        "Entity_Name",
        "Chunk_Type",
        "Char_Count",
        "Token_Estimate",
        "Embedding_Model",
        "Embedding_Dimension",
        "Created_Date",
        "Last_Updated",
        "Version",
        "Status",
        "Quality_Score",
        "Retrieval_Count",
        "Avg_Relevance_Score",
        "Notes",
        "Chunk_Text",
    ]
    chunk_rows: list[dict[str, Any]] = []
    for row in rows:
        document_id = as_text(row.get("Candidate_ID"))
        category, _ = subtopic_category_mapping(as_text(row.get("Primary_Subtopic")))
        chunks = row.get("_top_chunks") or []
        for idx, chunk in enumerate(chunks, start=1):
            text = as_text(chunk.get("text"))
            if not text:
                continue
            chunk_rows.append(
                {
                    "Chunk_ID": f"{document_id}_CHUNK_{idx:02d}",
                    "Document_ID": document_id,
                    "Category": category,
                    "Entity_Type": "Facet",
                    "Entity_Name": as_text(chunk.get("best_facet")) or as_text(row.get("Primary_Subtopic")),
                    "Chunk_Type": as_text(chunk.get("label")) or "evidence",
                    "Char_Count": len(text),
                    "Token_Estimate": max(1, math.ceil(len(text.split()) * 1.0)),
                    "Embedding_Model": as_text(row.get("Embedding_Model")) or as_text(row.get("Embedding_Backend")) or "hashing_bow",
                    "Embedding_Dimension": int(safe_float(row.get("Embedding_Dimension"), 512.0)),
                    "Created_Date": today,
                    "Last_Updated": today,
                    "Version": 1,
                    "Status": as_text(row.get("Final_Decision")).title() or "Curated",
                    "Quality_Score": round(safe_float(chunk.get("combined_score")), 2),
                    "Retrieval_Count": 0,
                    "Avg_Relevance_Score": round(safe_float(chunk.get("combined_score")), 2),
                    "Notes": normalize_space(
                        " ".join(
                            [
                                f"page={chunk.get('page_number')}" if chunk.get("page_number") else "",
                                f"facet={chunk.get('best_facet')}" if chunk.get("best_facet") else "",
                                text[:240],
                            ]
                        )
                    ),
                    "Chunk_Text": text,
                }
            )
    return pd.DataFrame(chunk_rows, columns=columns)


def processing_log_dataframe(rows: list[dict[str, Any]]) -> pd.DataFrame:
    today = datetime.now().strftime("%Y-%m-%d")
    columns = [
        "Log_ID",
        "Document_ID",
        "Document_Name",
        "Stage_ID",
        "Stage_Name",
        "Status",
        "Start_Time",
        "End_Time",
        "Duration_Minutes",
        "Records_Processed",
        "Records_Failed",
        "Error_Message",
        "Processed_By",
        "Notes",
    ]
    logs: list[dict[str, Any]] = []
    stages = [
        ("STAGE2", "Acquisition", "Acquisition_Status"),
        ("STAGE3", "Lightweight Card", "Stage3_Status"),
        ("STAGE6", "Enrichment", "Stage6_Status"),
        ("STAGE7", "LLM Judge", "LLM_Judge_Status"),
        ("FINAL", "Final Decision", "Final_Decision"),
    ]
    for row in rows:
        for stage_id, stage_name, status_field in stages:
            status_value = as_text(row.get(status_field))
            if not status_value:
                continue
            logs.append(
                {
                    "Log_ID": f"{as_text(row.get('Candidate_ID'))}_{stage_id}",
                    "Document_ID": as_text(row.get("Candidate_ID")),
                    "Document_Name": Path(as_text(row.get("Curated_File_Path")) or as_text(row.get("Document_File_Path"))).name
                    if as_text(row.get("Curated_File_Path")) or as_text(row.get("Document_File_Path"))
                    else sanitize_filename(as_text(row.get("Title"))),
                    "Stage_ID": stage_id,
                    "Stage_Name": stage_name,
                    "Status": status_value,
                    "Start_Time": "",
                    "End_Time": today,
                    "Duration_Minutes": "",
                    "Records_Processed": 1,
                    "Records_Failed": 0,
                    "Error_Message": status_value if "failed" in status_value.lower() else "",
                    "Processed_By": "gnem_pipeline",
                    "Notes": "",
                }
            )
    return pd.DataFrame(logs, columns=columns)


def load_grounding_companies_dataframe(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    try:
        excel = pd.ExcelFile(path)
        sheet_name = "Data" if "Data" in excel.sheet_names else excel.sheet_names[0]
        return pd.read_excel(path, sheet_name=sheet_name)
    except Exception:
        return pd.DataFrame()


def sqlite_safe_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    sql_df = df.copy()
    sql_df.columns = [re.sub(r"[^a-z0-9]+", "_", str(column).lower()).strip("_") for column in sql_df.columns]
    if hasattr(sql_df, "map"):
        sql_df = sql_df.map(lambda value: json.dumps(value, ensure_ascii=True) if isinstance(value, (dict, list)) else value)
    else:
        sql_df = sql_df.apply(
            lambda col: col.map(lambda value: json.dumps(value, ensure_ascii=True) if isinstance(value, (dict, list)) else value)
        )
    return sql_df


def write_sqlite_registry(
    path: Path,
    *,
    documents_df: pd.DataFrame,
    chunks_df: pd.DataFrame,
    grounding_companies_df: pd.DataFrame,
) -> None:
    if documents_df.columns.empty:
        documents_df = pd.DataFrame(columns=["Document_ID"])
    if chunks_df.columns.empty:
        chunks_df = pd.DataFrame(columns=["Chunk_ID"])
    if grounding_companies_df.columns.empty:
        grounding_companies_df = pd.DataFrame(columns=["Company"])
    with sqlite3.connect(path) as conn:
        documents_sql = sqlite_safe_dataframe(documents_df)
        chunks_sql = sqlite_safe_dataframe(chunks_df)
        grounding_sql = sqlite_safe_dataframe(grounding_companies_df)
        documents_sql.to_sql("documents", conn, if_exists="replace", index=False)
        chunks_sql.to_sql("chunks", conn, if_exists="replace", index=False)
        grounding_sql.to_sql("grounding_companies", conn, if_exists="replace", index=False)
        if "document_id" in documents_sql.columns:
            conn.execute("CREATE INDEX IF NOT EXISTS idx_documents_document_id ON documents(document_id)")
        if "processing_status" in documents_sql.columns:
            conn.execute("CREATE INDEX IF NOT EXISTS idx_documents_processing_status ON documents(processing_status)")
        if "document_id" in chunks_sql.columns:
            conn.execute("CREATE INDEX IF NOT EXISTS idx_chunks_document_id ON chunks(document_id)")
        if "entity_name" in chunks_sql.columns:
            conn.execute("CREATE INDEX IF NOT EXISTS idx_chunks_entity_name ON chunks(entity_name)")


def copy_curated_documents(rows: list[dict[str, Any]], final_docs_dir: Path) -> list[dict[str, Any]]:
    copied_rows: list[dict[str, Any]] = []
    keep_dir = final_docs_dir / "keep"
    review_dir = final_docs_dir / "review"
    keep_dir.mkdir(parents=True, exist_ok=True)
    review_dir.mkdir(parents=True, exist_ok=True)
    for row in rows:
        working = dict(row)
        decision = as_text(working.get("Final_Decision"))
        if decision not in {"keep", "review"}:
            working["Curated_Copy_Status"] = "not_curated"
            working["Curated_File_Path"] = ""
            copied_rows.append(working)
            continue
        source_path = (
            as_text(working.get("Document_File_Path"))
            or as_text(working.get("Acquired_File_Path"))
            or as_text(working.get("Resolved_File_Path"))
        )
        if not source_path or not Path(source_path).exists():
            working["Curated_Copy_Status"] = "missing_source"
            working["Curated_File_Path"] = ""
            copied_rows.append(working)
            continue
        target_dir = keep_dir if decision == "keep" else review_dir
        source = Path(source_path)
        target_name = safe_output_filename(target_dir, f"{as_text(working.get('Candidate_ID'))}_{source.name}")
        target = target_dir / target_name
        suffix = source.suffix
        stem = target.stem
        attempt = 2
        while target.exists():
            target = target_dir / safe_output_filename(target_dir, f"{stem}_{attempt}{suffix}")
            attempt += 1
        shutil.copy2(source, target)
        working["Curated_Copy_Status"] = "copied"
        working["Curated_File_Path"] = str(target.resolve())
        copied_rows.append(working)
    return copied_rows


def cleanup_stage_downloads(download_dir: Path, *, keep_stage_downloads: bool) -> tuple[str, int]:
    if keep_stage_downloads:
        return "kept_by_flag", 0
    if not download_dir.exists():
        return "not_found", 0
    try:
        file_count = sum(1 for path in download_dir.rglob("*") if path.is_file())
        shutil.rmtree(download_dir)
        return "removed", file_count
    except Exception as exc:
        return f"cleanup_failed:{exc}", 0


def run_rag_filtering_pipeline(
    *,
    args: argparse.Namespace,
    llm_cfg: LLMConfig | None,
    golden_summary: str,
    out_dir: Path,
    stage1_df: pd.DataFrame,
    stage1_dedup_df: pd.DataFrame,
) -> int:
    stage_artifacts_enabled = bool(args.write_stage_artifacts)
    cwd = Path.cwd()
    grounding_xlsx = resolve_existing_repo_path(
        args.grounding_xlsx,
        DEFAULT_GROUNDING_XLSX,
        LEGACY_GROUNDING_XLSX,
        ALT_LEGACY_GROUNDING_XLSX,
    ) or Path(args.grounding_xlsx)
    grounding_geojson = resolve_existing_repo_path(
        args.grounding_counties_geojson,
        DEFAULT_GROUNDING_GEOJSON,
        LEGACY_GROUNDING_GEOJSON,
    ) or Path(args.grounding_counties_geojson)
    grounding_docx = resolve_existing_repo_path(
        args.grounding_docx,
        DEFAULT_GROUNDING_DOCX,
        LEGACY_GROUNDING_DOCX,
    ) or Path(args.grounding_docx)

    grounding = build_grounding_dictionaries(
        automotive_xlsx=grounding_xlsx,
        counties_geojson=grounding_geojson,
        supply_chain_docx=grounding_docx,
        golden_summary=golden_summary,
    )
    grounding_path = out_dir / "grounding_summary.json"
    grounding_path.write_text(json.dumps(grounding_summary_payload(grounding), indent=2), encoding="utf-8")
    print(f"Grounding summary: {grounding_path}")
    if not grounding.companies and not grounding.counties and not grounding.oems:
        print("Warning: grounding dictionaries are empty for companies/counties/OEMs; retrieval quality may degrade.")

    embedding_runtime: EmbeddingRuntime | None = None
    if args.embedding_provider != "none":
        embedding_base_url = args.embedding_base_url
        if args.embedding_provider == "openai" and embedding_base_url.strip() == "http://localhost:11434":
            embedding_base_url = "https://api.openai.com"
        runtime_candidate = EmbeddingRuntime(
            EmbeddingConfig(
                provider=args.embedding_provider,
                model=args.embedding_model,
                base_url=embedding_base_url,
                api_key=args.embedding_api_key,
                timeout_sec=max(10, int(args.embedding_timeout)),
                batch_size=max(1, int(args.embedding_batch_size)),
            )
        )
        warmup_texts = [golden_summary, grounding.global_reference, *list(grounding.facet_texts.values())[:2]]
        if runtime_candidate.warmup(warmup_texts):
            embedding_runtime = runtime_candidate
            print(f"Embedding mode: provider={args.embedding_provider}, model={args.embedding_model}, backend={runtime_candidate.backend_name}")
        else:
            print(f"Embedding mode: fallback to hashing_bow ({runtime_candidate.last_error or 'embedding_warmup_failed'})")
    else:
        print("Embedding mode: hashing_bow fallback only")

    auto_doc_dirs = [cwd / DEFAULT_LOCAL_DOC_DIR, cwd / LEGACY_LOCAL_DOC_DIR]
    auto_text_dirs = [cwd / DEFAULT_LOCAL_TEXT_DIR, cwd / LEGACY_LOCAL_TEXT_DIR]
    local_doc_dirs = [Path(p) for p in args.local_pdf_dir if p]
    local_text_dirs = [Path(p) for p in args.local_text_dir if p]
    if not local_doc_dirs:
        local_doc_dirs = [p for p in auto_doc_dirs if p.exists()]
    if not local_text_dirs:
        local_text_dirs = [p for p in auto_text_dirs if p.exists()]
    local_doc_index = build_local_document_index(local_doc_dirs)
    local_text_index = build_local_text_index(local_text_dirs)
    print(f"Local document index size: {len(local_doc_index)}")
    print(f"Local text index size: {len(local_text_index)}")

    s2 = stage1_dedup_df.copy()
    if "Specificity_Hits" not in s2.columns:
        s2["Specificity_Hits"] = 0
    if "Metadata_TwoSignal_Rule" not in s2.columns:
        s2["Metadata_TwoSignal_Rule"] = (s2["Signal_Category_Count"].fillna(0).astype(float) >= 2).astype(int)
    if "Metadata_QuestionSpecifics_Rule" not in s2.columns:
        s2["Metadata_QuestionSpecifics_Rule"] = (
            (s2["Question_Coverage"].fillna(0).astype(float) >= 1)
            & (s2["Specificity_Hits"].fillna(0).astype(float) >= 1)
        ).astype(int)
    s2["Metadata_Rule_Pass"] = (
        (s2["Metadata_TwoSignal_Rule"].astype(int) == 1)
        | (s2["Metadata_QuestionSpecifics_Rule"].astype(int) == 1)
    ).astype(int)
    s2["Metadata_Pass"] = (s2["Metadata_Score"] >= float(args.metadata_threshold)).astype(int)

    resolved_rows: list[dict[str, Any]] = []
    for row in s2.to_dict(orient="records"):
        working = dict(row)
        working.update(resolve_document_paths(working, local_doc_index, local_text_index))
        working["Document_Candidate"] = 1 if probable_document(working) else 0
        working["Local_Path_Found"] = 1 if as_text(working.get("Resolved_File_Path")) else 0
        working["Text_Path_Found"] = 1 if as_text(working.get("Resolved_Text_Path")) else 0
        resolved_rows.append(working)
    s2 = pd.DataFrame(resolved_rows)

    if args.disable_metadata_rule_gate:
        s2["Stage2_Filter_Pass"] = ((s2["Metadata_Pass"] == 1) & (s2["Document_Candidate"] == 1)).astype(int)
    else:
        s2["Stage2_Filter_Pass"] = (
            (s2["Metadata_Pass"] == 1) & (s2["Metadata_Rule_Pass"] == 1) & (s2["Document_Candidate"] == 1)
        ).astype(int)
    s2 = s2.sort_values(by=["Metadata_Score"], ascending=False).reset_index(drop=True)
    stage2_pool_path = out_dir / "stage2_scored_pool.xlsx"
    maybe_write_excel(stage_artifacts_enabled, s2, stage2_pool_path)
    print(f"Stage 2 scored pool: {len(s2)}" + (f" -> {stage2_pool_path}" if stage_artifacts_enabled else ""))

    selected_for_acquisition = s2[s2["Stage2_Filter_Pass"] == 1].copy()
    selected_for_acquisition = selected_for_acquisition.sort_values(by=["Metadata_Score"], ascending=False).reset_index(drop=True)
    if args.metadata_target_ratio and args.metadata_target_ratio > 0:
        target_n = max(1, math.ceil(len(stage1_dedup_df) * float(args.metadata_target_ratio)))
        if len(selected_for_acquisition) > target_n:
            selected_for_acquisition = selected_for_acquisition.head(target_n).copy()
        print(f"Stage 2 ratio cap: {target_n} (ratio={args.metadata_target_ratio})")
    if args.sample_size and args.sample_size > 0:
        selected_for_acquisition = selected_for_acquisition.head(args.sample_size).copy()

    download_dir = out_dir / "stage2_downloads"
    acquired_rows: list[dict[str, Any]] = []
    for i, row in enumerate(selected_for_acquisition.to_dict(orient="records"), start=1):
        if i % 25 == 0 or i == 1:
            print(f"[stage2 {i}/{len(selected_for_acquisition)}] acquiring document")
        working = dict(row)
        working["Acquisition_Status"] = "NotAcquired"
        working["Acquired_File_Path"] = ""
        working["Acquired_Size_MB"] = 0.0
        working["Download_Content_Type"] = ""

        if as_text(working.get("Resolved_File_Path")):
            working["Acquisition_Status"] = "LocalFound"
            working["Acquired_File_Path"] = as_text(working.get("Resolved_File_Path"))
            try:
                working["Acquired_Size_MB"] = round(Path(working["Acquired_File_Path"]).stat().st_size / (1024 * 1024), 2)
            except OSError:
                working["Acquired_Size_MB"] = 0.0
            acquired_rows.append(working)
            continue

        url = as_text(working.get("URL"))
        if not url:
            acquired_rows.append(working)
            continue

        fallback_name = sanitize_filename(
            as_text(working.get("Filename")) or as_text(working.get("Title")) or as_text(working.get("Candidate_ID")) or "candidate"
        )
        dl_path, dl_mb, dl_status, content_type = download_document(
            url=url,
            output_dir=download_dir,
            fallback_name=fallback_name,
            timeout_sec=args.timeout,
            max_download_mb=args.max_download_mb,
        )
        working["Acquisition_Status"] = dl_status
        working["Acquired_File_Path"] = dl_path
        working["Acquired_Size_MB"] = dl_mb
        working["Download_Content_Type"] = content_type
        acquired_rows.append(working)

    stage2_df = pd.DataFrame(acquired_rows)
    stage2_selected_path = out_dir / "stage2_selected_by_metadata.xlsx"
    maybe_write_excel(stage_artifacts_enabled, stage2_df, stage2_selected_path)
    print(f"Stage 2 selected rows: {len(stage2_df)}" + (f" -> {stage2_selected_path}" if stage_artifacts_enabled else ""))

    acq_lookup = {as_text(row.get("Candidate_ID")): row for row in stage2_df.to_dict(orient="records")}
    stage2_all_rows: list[dict[str, Any]] = []
    for row in s2.to_dict(orient="records"):
        working = dict(row)
        extra = acq_lookup.get(as_text(working.get("Candidate_ID")), {})
        if extra:
            working.update(extra)
        if not as_text(working.get("Acquired_File_Path")) and as_text(working.get("Resolved_File_Path")):
            working["Acquired_File_Path"] = as_text(working.get("Resolved_File_Path"))
            working["Acquisition_Status"] = as_text(working.get("Acquisition_Status")) or "LocalResolved"
        populate_source_file_fields(working)
        stage2_all_rows.append(working)

    stage2_all_df = pd.DataFrame(stage2_all_rows)
    stage2_ready_path = out_dir / "stage2_acquired_documents.xlsx"
    maybe_write_excel(stage_artifacts_enabled, stage2_all_df, stage2_ready_path)
    print(f"Stage 2 resolved/acquired documents: {len(stage2_all_df)}" + (f" -> {stage2_ready_path}" if stage_artifacts_enabled else ""))
    stage2_pdf_path = out_dir / "stage2_acquired_pdfs.xlsx"
    maybe_write_excel(
        stage_artifacts_enabled,
        stage2_all_df[
            stage2_all_df["Acquired_File_Path"].astype(str).str.lower().str.endswith(".pdf")
        ],
        stage2_pdf_path,
    )

    hybrid_threshold = float(args.hybrid_threshold if args.hybrid_threshold is not None else args.embedding_threshold)
    credibility_threshold = float(args.credibility_threshold)
    effective_hybrid_threshold = hybrid_threshold
    effective_direct_usecase_threshold = float(args.direct_usecase_threshold)
    embedding_backend_name = embedding_runtime.backend_name if embedding_runtime else "hashing_bow"
    if embedding_backend_name == "hashing_bow":
        effective_hybrid_threshold = min(effective_hybrid_threshold, 50.0)
        effective_direct_usecase_threshold = min(effective_direct_usecase_threshold, 0.55)
        print(
            "Embedding fallback active: using relaxed shortlist thresholds "
            f"(hybrid={effective_hybrid_threshold}, direct_usecase={effective_direct_usecase_threshold})."
        )

    card_df = stage2_all_df[
        (stage2_all_df["Stage2_Filter_Pass"] == 1)
        & (stage2_all_df["Document_Candidate"] == 1)
        & (stage2_all_df["Source_File_Available"] == 1)
    ].copy()
    card_df = card_df.sort_values(by=["Metadata_Score"], ascending=False).reset_index(drop=True)
    if args.sample_size and args.sample_size > 0:
        card_df = card_df.head(args.sample_size).copy()
        print(f"Sample size applied to document-card pool: {len(card_df)}")

    card_rows: list[dict[str, Any]] = []
    for i, row in enumerate(card_df.to_dict(orient="records"), start=1):
        if i % 25 == 0 or i == 1:
            print(f"[stage3 {i}/{len(card_df)}] building lightweight document cards")
        working = dict(row)
        working["_similarity_runtime"] = embedding_runtime
        try:
            working.update(
                build_document_card(
                    working,
                    grounding,
                    max_text_chars=max(5000, int(args.document_max_chars)),
                    card_level="lightweight",
                    similarity_runtime=embedding_runtime,
                )
            )
            working["Stage3_Status"] = "LightweightCardBuilt"
        except Exception as exc:
            working["Stage3_Status"] = f"CardFailed:{exc}"
            working["Card_Level"] = "lightweight"
            working["_card_text"] = normalize_space(
                " ".join([as_text(working.get("Title")), as_text(working.get("Description")), as_text(working.get("Content_Snippet"))])
            )
            working["_top_chunks"] = []
        card_rows.append(working)

    stage3_cards_path = out_dir / "stage3_document_cards.xlsx"
    stage3_lightweight_path = out_dir / "stage3_lightweight_document_cards.xlsx"
    stage3_cards_jsonl = out_dir / "stage3_document_cards.jsonl"
    stage3_lightweight_jsonl = out_dir / "stage3_lightweight_document_cards.jsonl"
    stage3_export_rows = [{k: v for k, v in row.items() if not str(k).startswith("_")} for row in card_rows]
    stage3_df = pd.DataFrame(stage3_export_rows)
    maybe_write_excel(stage_artifacts_enabled, stage3_df, stage3_cards_path)
    maybe_write_excel(stage_artifacts_enabled, stage3_df, stage3_lightweight_path)
    maybe_write_jsonl(stage_artifacts_enabled, stage3_cards_jsonl, card_rows)
    maybe_write_jsonl(stage_artifacts_enabled, stage3_lightweight_jsonl, card_rows)
    print(f"Stage 3 lightweight document cards: {len(stage3_export_rows)}" + (f" -> {stage3_lightweight_path}" if stage_artifacts_enabled else ""))

    for row in card_rows:
        row.update(score_document_card(row, grounding, float(row.get("Metadata_Score", 0.0))))
        row["Heuristic_Pass"] = 1 if float(row.get("Heuristic_Score", 0.0)) >= float(args.heuristic_threshold) else 0
        row["Embedding_Pass"] = 1 if float(row.get("Embedding_Score", 0.0)) >= float(args.embedding_threshold) else 0
        row["Hybrid_Pass"] = 1 if float(row.get("Hybrid_Score", 0.0)) >= effective_hybrid_threshold else 0
    stage4_path = out_dir / "stage4_embedding_scored.xlsx"
    stage4_hybrid_path = out_dir / "stage4_hybrid_scored.xlsx"
    stage4_df = pd.DataFrame([{k: v for k, v in row.items() if not str(k).startswith("_")} for row in card_rows])
    maybe_write_excel(stage_artifacts_enabled, stage4_df, stage4_path)
    maybe_write_excel(stage_artifacts_enabled, stage4_df, stage4_hybrid_path)
    print(f"Stage 4 hybrid scored: {len(card_rows)}" + (f" -> {stage4_hybrid_path}" if stage_artifacts_enabled else ""))

    for row in card_rows:
        row.update(classify_document_card(row, row))
        preliminary_shortlist_pass = (
            float(row.get("Heuristic_Score", 0.0)) >= float(args.heuristic_threshold)
            and float(row.get("Hybrid_Score", 0.0)) >= effective_hybrid_threshold
            and float(row.get("Direct_Usecase_Score", 0.0)) >= effective_direct_usecase_threshold
        )
        row["Preliminary_Shortlist_Pass"] = 1 if preliminary_shortlist_pass else 0
        row["Shortlist_Pass"] = 1 if preliminary_shortlist_pass else 0
        row["Shortlist_Selected_By_Ratio"] = 0
        row["Promoted_To_Review_Flag"] = 0
    stage5_path = out_dir / "stage5_classifier_reranked.xlsx"
    stage5_df = pd.DataFrame([{k: v for k, v in row.items() if not str(k).startswith("_")} for row in card_rows])
    maybe_write_excel(stage_artifacts_enabled, stage5_df, stage5_path)
    print(f"Stage 5 classifier scored: {len(card_rows)}" + (f" -> {stage5_path}" if stage_artifacts_enabled else ""))

    shortlist_seed_rows = [row for row in card_rows if int(row.get("Preliminary_Shortlist_Pass", 0)) == 1]
    enriched_rows: list[dict[str, Any]] = []
    enriched_lookup: dict[str, dict[str, Any]] = {}
    enriched_max_chars = max(max(5000, int(args.document_max_chars)), 120000)
    for i, row in enumerate(shortlist_seed_rows, start=1):
        if i % 10 == 0 or i == 1:
            print(f"[stage6 {i}/{len(shortlist_seed_rows)}] building enriched cards")
        working = dict(row)
        working["_similarity_runtime"] = embedding_runtime
        try:
            working.update(
                build_document_card(
                    working,
                    grounding,
                    max_text_chars=enriched_max_chars,
                    card_level="enriched",
                    similarity_runtime=embedding_runtime,
                )
            )
            working.update(score_document_card(working, grounding, float(working.get("Metadata_Score", 0.0))))
            working.update(classify_document_card(working, working))
            working["Stage6_Status"] = "EnrichedCardBuilt"
        except Exception as exc:
            working["Stage6_Status"] = f"EnrichmentFailed:{exc}"
        enriched_rows.append(working)
        enriched_lookup[as_text(working.get("Candidate_ID"))] = working

    for idx, row in enumerate(card_rows):
        enriched = enriched_lookup.get(as_text(row.get("Candidate_ID")))
        if enriched:
            card_rows[idx] = enriched

    card_rows = apply_exact_duplicate_pass(card_rows)
    card_rows = apply_near_duplicate_pass(card_rows)

    ranked_rows = sorted(
        [
            row
            for row in card_rows
            if int(row.get("Preliminary_Shortlist_Pass", 0)) == 1
            and int(row.get("Duplicate_Master_Flag", 1)) == 1
            and int(row.get("Near_Duplicate_Master_Flag", 1)) == 1
        ],
        key=lambda item: float(item.get("Rerank_Score", 0.0)),
        reverse=True,
    )
    shortlist_rows = list(ranked_rows)
    if args.final_target_ratio and args.final_target_ratio > 0 and ranked_rows:
        shortlist_cap = max(1, math.ceil(len(ranked_rows) * float(args.final_target_ratio)))
        shortlist_rows = shortlist_rows[:shortlist_cap]
        print(f"Shortlist ratio cap: {shortlist_cap} (ratio={args.final_target_ratio})")
    shortlist_ids = {as_text(row.get("Candidate_ID")) for row in shortlist_rows}
    for row in card_rows:
        row["Shortlist_Pass"] = 1 if as_text(row.get("Candidate_ID")) in shortlist_ids else 0
        row["Shortlist_Selected_By_Ratio"] = 1 if as_text(row.get("Candidate_ID")) in shortlist_ids else 0

    stage6_path = out_dir / "stage6_enriched_document_cards.xlsx"
    stage6_jsonl = out_dir / "stage6_enriched_document_cards.jsonl"
    stage6_df = pd.DataFrame([{k: v for k, v in row.items() if not str(k).startswith("_")} for row in enriched_rows])
    maybe_write_excel(stage_artifacts_enabled, stage6_df, stage6_path)
    maybe_write_jsonl(stage_artifacts_enabled, stage6_jsonl, enriched_rows)
    print(f"Stage 6 enriched cards: {len(enriched_rows)}" + (f" -> {stage6_path}" if stage_artifacts_enabled else ""))

    for row in card_rows:
        row["LLM_Judge_Status"] = "not_used"
        row["LLM_Relevance_Score"] = ""
        row["LLM_Usecase_Match"] = ""
        row["LLM_Information_Quality"] = ""
        row["LLM_Noise_Level"] = ""
        row["LLM_Model_Decision"] = ""
        row["LLM_Decision"] = ""
        row["LLM_Reason"] = ""
        row["LLM_Top_Evidence_Used"] = ""
        row["LLM_Confidence"] = 0.0
        row["LLM_Judge_Pass"] = 0
    if llm_cfg:
        for i, row in enumerate(shortlist_rows, start=1):
            if i % 10 == 0 or i == 1:
                print(f"[stage7 {i}/{len(shortlist_rows)}] llm judge")
            try:
                row.update(llm_document_judge(cfg=llm_cfg, golden_summary=golden_summary, card_row=row))
            except Exception as exc:
                row["LLM_Judge_Status"] = f"failed:{exc}"

    stage7_path = out_dir / "stage7_llm_judged.xlsx"
    stage7_legacy_path = out_dir / "stage6_llm_judged.xlsx"
    stage7_df = pd.DataFrame([{k: v for k, v in row.items() if not str(k).startswith("_")} for row in card_rows])
    maybe_write_excel(stage_artifacts_enabled, stage7_df, stage7_path)
    maybe_write_excel(stage_artifacts_enabled, stage7_df, stage7_legacy_path)
    print(f"Stage 7 llm judged: {len(card_rows)}" + (f" -> {stage7_path}" if stage_artifacts_enabled else ""))

    for row in card_rows:
        row.update(assess_document_credibility(row))
        row["Credibility_Pass"] = 1 if float(row.get("Credibility_Score", 0.0)) >= credibility_threshold else 0
        row["Final_Rank_Score"] = round(
            0.40 * float(row.get("Rerank_Score", 0.0))
            + 0.25 * float(row.get("Hybrid_Score", 0.0))
            + 0.25 * float(row.get("Credibility_Score", 0.0))
            + 5.0 * float(row.get("LLM_Judge_Pass", 0.0))
            + 0.5 * float(row.get("LLM_Information_Quality", 0.0) or 0.0),
            2,
        )
    stage8_path = out_dir / "stage8_credibility_checked.xlsx"
    stage8_df = pd.DataFrame([{k: v for k, v in row.items() if not str(k).startswith("_")} for row in card_rows])
    maybe_write_excel(stage_artifacts_enabled, stage8_df, stage8_path)
    print(f"Stage 8 credibility checked: {len(card_rows)}" + (f" -> {stage8_path}" if stage_artifacts_enabled else ""))

    decision_order = {"keep": 0, "review": 1, "discard": 2}
    for row in card_rows:
        if int(row.get("Exact_Duplicate_Flag", 0)) == 1 and int(row.get("Duplicate_Master_Flag", 1)) == 0:
            decision = "discard"
            reason = (
                f"Exact duplicate of {as_text(row.get('Duplicate_Of_Candidate_ID'))} based on "
                f"{as_text(row.get('Duplicate_Group_Type'))}."
            )
        elif int(row.get("Near_Duplicate_Flag", 0)) == 1 and int(row.get("Near_Duplicate_Master_Flag", 1)) == 0:
            decision = "discard"
            reason = (
                f"Near duplicate of {as_text(row.get('Near_Duplicate_Of_Candidate_ID'))} "
                f"(hamming={as_text(row.get('Near_Duplicate_Min_Hamming'))}, "
                f"text_similarity={as_text(row.get('Near_Duplicate_Text_Similarity'))})."
            )
        else:
            decision, reason = final_decision_reason(row, llm_enabled=bool(llm_cfg))
        row["Final_Decision"] = decision
        row["Keep_Review_Discard_Reason"] = reason
        row["Source"] = as_text(row.get("Source_Domain")) or as_text(row.get("Source"))
        row["File_Path_Or_URL"] = (
            as_text(row.get("Document_File_Path"))
            or as_text(row.get("Document_URL"))
            or as_text(row.get("Document_Source"))
        )
    scored_final_rows = sorted(
        card_rows,
        key=lambda item: (
            decision_order.get(as_text(item.get("Final_Decision")), 9),
            -float(item.get("Final_Rank_Score", 0.0)),
            -float(item.get("Hybrid_Score", 0.0)),
            -float(item.get("Credibility_Score", 0.0)),
        ),
    )
    scored_final_rows = apply_diversity_pass(scored_final_rows, max_share=float(args.diversity_max_share))
    scored_final_rows = sorted(
        scored_final_rows,
        key=lambda item: (
            decision_order.get(as_text(item.get("Final_Decision")), 9),
            -float(item.get("Final_Rank_Score", 0.0)),
            -float(item.get("Hybrid_Score", 0.0)),
            -float(item.get("Credibility_Score", 0.0)),
        ),
    )
    target_curated_min = int(args.curated_min_count) if len(stage1_df) >= int(args.curated_band_min_stage1_count) else 0
    target_curated_max = int(args.curated_max_count) if len(stage1_df) >= int(args.curated_band_min_stage1_count) else 0
    if target_curated_max > 0 and target_curated_max < target_curated_min:
        target_curated_max = target_curated_min
    promoted_to_review_count = 0
    current_curated_count = sum(1 for row in scored_final_rows if as_text(row.get("Final_Decision")) in {"keep", "review"})
    if target_curated_min > 0 and current_curated_count < target_curated_min:
        promotion_slots = target_curated_min - current_curated_count
        if target_curated_max > 0:
            promotion_slots = min(promotion_slots, max(0, target_curated_max - current_curated_count))
        promotion_candidates = sorted(
            [
                row for row in scored_final_rows
                if eligible_for_review_promotion(
                    row,
                    credibility_threshold=credibility_threshold,
                    heuristic_threshold=float(args.heuristic_threshold),
                    hybrid_threshold=effective_hybrid_threshold,
                    direct_usecase_threshold=effective_direct_usecase_threshold,
                )
            ],
            key=final_rank_sort_key,
            reverse=True,
        )
        for row in promotion_candidates[:promotion_slots]:
            existing_reason = as_text(row.get("Keep_Review_Discard_Reason"))
            row["Final_Decision"] = "review"
            row["Promoted_To_Review_Flag"] = 1
            row["Keep_Review_Discard_Reason"] = normalize_space(
                f"{existing_reason} Promoted to review to meet the large-run curated recall band after clearing secondary quality checks."
            )
            promoted_to_review_count += 1

    scored_final_rows = sorted(
        scored_final_rows,
        key=lambda item: (
            decision_order.get(as_text(item.get("Final_Decision")), 9),
            -float(item.get("Final_Rank_Score", 0.0)),
            -float(item.get("Hybrid_Score", 0.0)),
            -float(item.get("Credibility_Score", 0.0)),
        ),
    )

    attempted_curated_rows = copy_curated_documents(
        [row for row in scored_final_rows if as_text(row.get("Final_Decision")) in {"keep", "review"}],
        out_dir / "final_docs",
    )
    attempted_lookup = {as_text(row.get("Candidate_ID")): row for row in attempted_curated_rows}
    for idx, row in enumerate(scored_final_rows):
        merged = attempted_lookup.get(as_text(row.get("Candidate_ID")))
        if merged:
            populate_source_file_fields(merged)
            if as_text(merged.get("Final_Decision")) in {"keep", "review"} and as_text(merged.get("Curated_Copy_Status")) != "copied":
                merged["Final_Decision"] = "discard"
                merged["Keep_Review_Discard_Reason"] = "Rejected because the source file was not preserved locally for final curation."
                merged["Curated_File_Path"] = ""
            scored_final_rows[idx] = merged
        else:
            scored_final_rows[idx]["Curated_Copy_Status"] = "not_curated"
            scored_final_rows[idx]["Curated_File_Path"] = ""
            populate_source_file_fields(scored_final_rows[idx])

    scored_lookup = {as_text(row.get("Candidate_ID")): row for row in scored_final_rows}
    all_final_rows: list[dict[str, Any]] = []
    for row in stage2_all_rows:
        candidate_id = as_text(row.get("Candidate_ID"))
        working = dict(scored_lookup.get(candidate_id, dict(row)))
        populate_source_file_fields(working)
        if candidate_id not in scored_lookup:
            working["Final_Decision"] = "discard"
            working["Keep_Review_Discard_Reason"] = stage2_rejection_reason(working)
            working["Curated_Copy_Status"] = "not_curated"
            working["Curated_File_Path"] = ""
            working["Final_Rank_Score"] = round(safe_float(working.get("Metadata_Score", 0.0)), 2)
            working["Shortlist_Pass"] = 0
            working["Promoted_To_Review_Flag"] = 0
        working["Source"] = as_text(working.get("Source_Domain")) or as_text(working.get("Source"))
        working["File_Path_Or_URL"] = (
            existing_file_path(
                working.get("Curated_File_Path"),
                working.get("Document_File_Path"),
                working.get("Acquired_File_Path"),
                working.get("Resolved_File_Path"),
            )
            or as_text(working.get("Document_URL"))
            or as_text(working.get("URL"))
            or as_text(working.get("Document_Source"))
        )
        all_final_rows.append(working)

    all_final_rows = sorted(
        all_final_rows,
        key=lambda item: (
            decision_order.get(as_text(item.get("Final_Decision")), 9),
            -float(item.get("Final_Rank_Score", 0.0)),
            -float(item.get("Hybrid_Score", 0.0)),
            -float(item.get("Credibility_Score", 0.0)),
        ),
    )

    curated_rows = [row for row in all_final_rows if as_text(row.get("Final_Decision")) in {"keep", "review"}]
    rejected_rows = [row for row in all_final_rows if as_text(row.get("Final_Decision")) == "discard"]
    curated_export_rows = [{k: v for k, v in row.items() if not str(k).startswith("_")} for row in curated_rows]
    rejected_export_rows = [{k: v for k, v in row.items() if not str(k).startswith("_")} for row in rejected_rows]
    final_df = ensure_columns(
        pd.DataFrame(curated_export_rows) if curated_export_rows else pd.DataFrame(columns=FINAL_EXPORT_COLUMNS),
        preferred=FINAL_EXPORT_COLUMNS,
    )
    rejected_df = ensure_columns(
        pd.DataFrame(rejected_export_rows) if rejected_export_rows else pd.DataFrame(columns=REJECTED_EXPORT_COLUMNS),
        preferred=REJECTED_EXPORT_COLUMNS,
    )
    final_path = out_dir / "review_ready_documents.xlsx"
    rejected_path = out_dir / "rejected_documents.xlsx"
    final_csv_path = out_dir / "review_ready_documents.csv"
    rejected_csv_path = out_dir / "rejected_documents.csv"
    final_jsonl_path = out_dir / "review_ready_documents.jsonl"
    rejected_jsonl_path = out_dir / "rejected_documents.jsonl"
    curated_jsonl_path = out_dir / "curated_documents.jsonl"
    write_excel(final_df, final_path)
    write_excel(rejected_df, rejected_path)
    write_jsonl(final_jsonl_path, curated_rows)
    write_jsonl(rejected_jsonl_path, rejected_rows)
    write_jsonl(curated_jsonl_path, curated_rows)
    if args.write_csv_exports:
        final_df.to_csv(final_csv_path, index=False)
        rejected_df.to_csv(rejected_csv_path, index=False)
    print(f"Final review-ready documents: {len(final_df)} -> {final_path}")
    print(f"Rejected documents: {len(rejected_df)} -> {rejected_path}")

    curated_document_df = document_registry_dataframe(curated_rows)
    curated_chunk_df = chunk_registry_dataframe(curated_rows)
    processing_log_df = processing_log_dataframe(all_final_rows)
    registry_workbook_path = out_dir / "rag_data_management_registry.xlsx"
    write_excel_sheets(
        registry_workbook_path,
        {
            "Document_Registry": curated_document_df,
            "Chunk_Registry": curated_chunk_df,
            "Processing_Log": processing_log_df,
            "Review_Ready": final_df,
            "Rejected": rejected_df,
        },
    )
    write_excel(curated_document_df, out_dir / "document_registry.xlsx")
    write_excel(curated_chunk_df, out_dir / "chunk_registry.xlsx")
    if args.write_csv_exports:
        curated_document_df.to_csv(out_dir / "document_registry.csv", index=False)
        curated_chunk_df.to_csv(out_dir / "chunk_registry.csv", index=False)

    grounding_companies_df = load_grounding_companies_dataframe(grounding_xlsx)
    sqlite_path = out_dir / "rag_registry.sqlite"
    print(f"RAG registry workbook: {registry_workbook_path}")
    if args.write_sqlite_registry:
        write_sqlite_registry(
            sqlite_path,
            documents_df=curated_document_df,
            chunks_df=curated_chunk_df,
            grounding_companies_df=grounding_companies_df,
        )
        print(f"SQLite registry: {sqlite_path}")

    published_ready_manifest_rows: list[dict[str, Any]] = []
    published_ready_unmatched_rows: list[dict[str, Any]] = []
    published_ready_dir = ""
    if not args.skip_ready_docs_publish:
        ready_dir_str = as_text(args.publish_ready_docs_dir)
        if ready_dir_str:
            ready_dir = Path(ready_dir_str)
            published_ready_manifest_rows, published_ready_unmatched_rows = publish_curated_documents_to_ready_dir(
                curated_rows,
                ready_dir,
            )
            published_ready_dir = str(ready_dir.resolve())
            print(
                "Published curated Tavily documents for evAutomationUpdated: "
                f"{len(published_ready_manifest_rows)} files -> {published_ready_dir}"
            )

    cleanup_status, cleaned_file_count = cleanup_stage_downloads(
        download_dir,
        keep_stage_downloads=bool(args.keep_stage_downloads),
    )
    if cleanup_status == "removed":
        print(f"Stage 2 downloads cleaned up: removed {cleaned_file_count} file(s)")
    elif cleanup_status == "kept_by_flag":
        print("Stage 2 downloads kept by flag")
    elif cleanup_status != "not_found":
        print(f"Stage 2 download cleanup status: {cleanup_status}")

    report = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "output_dir": str(out_dir.resolve()),
        "stage1_all_candidates": int(len(stage1_df)),
        "stage1_dedup_candidates": int(len(stage1_dedup_df)),
        "stage2_scored_pool_count": int(len(s2)),
        "stage2_selected_by_metadata_count": int(len(stage2_df)),
        "stage2_document_pool_count": int(len(stage2_all_df)),
        "stage3_lightweight_card_count": int(len(card_rows)),
        "stage6_enriched_card_count": int(len(enriched_rows)),
        "stage7_shortlist_count": int(sum(1 for row in card_rows if int(row.get("Shortlist_Pass", 0)) == 1)),
        "stage7_llm_judged_count": int(sum(1 for row in card_rows if as_text(row.get("LLM_Judge_Status")) == "ok")),
        "stage8_credible_count": int(sum(1 for row in card_rows if int(row.get("Credibility_Pass", 0)) == 1)),
        "exact_duplicate_grouped_count": int(sum(1 for row in card_rows if int(row.get("Exact_Duplicate_Flag", 0)) == 1)),
        "near_duplicate_grouped_count": int(sum(1 for row in card_rows if int(row.get("Near_Duplicate_Flag", 0)) == 1)),
        "final_keep_count": int(sum(1 for row in all_final_rows if as_text(row.get("Final_Decision")) == "keep")),
        "final_review_count": int(sum(1 for row in all_final_rows if as_text(row.get("Final_Decision")) == "review")),
        "final_discard_count": int(sum(1 for row in all_final_rows if as_text(row.get("Final_Decision")) == "discard")),
        "final_curated_document_count": int(len(curated_document_df)),
        "final_rejected_document_count": int(len(rejected_rows)),
        "final_chunk_registry_count": int(len(curated_chunk_df)),
        "final_review_ready_jsonl": str(final_jsonl_path),
        "final_rejected_jsonl": str(rejected_jsonl_path),
        "final_curated_jsonl": str(curated_jsonl_path),
        "sqlite_registry_written": bool(args.write_sqlite_registry),
        "stage2_download_cleanup_status": cleanup_status,
        "stage2_download_cleanup_files_removed": int(cleaned_file_count),
        "stage_artifacts_written": bool(args.write_stage_artifacts),
        "csv_exports_written": bool(args.write_csv_exports),
        "promoted_to_review_count": int(promoted_to_review_count),
        "published_ready_docs_count": int(len(published_ready_manifest_rows)),
        "published_ready_docs_unmatched_count": int(len(published_ready_unmatched_rows)),
        "published_ready_docs_dir": published_ready_dir,
        "thresholds": {
            "metadata_threshold": float(args.metadata_threshold),
            "heuristic_threshold": float(args.heuristic_threshold),
            "embedding_threshold": float(args.embedding_threshold),
            "hybrid_threshold": hybrid_threshold,
            "effective_hybrid_threshold": effective_hybrid_threshold,
            "direct_usecase_threshold": float(args.direct_usecase_threshold),
            "effective_direct_usecase_threshold": effective_direct_usecase_threshold,
            "credibility_threshold": credibility_threshold,
            "metadata_target_ratio": float(args.metadata_target_ratio),
            "final_target_ratio": float(args.final_target_ratio),
            "diversity_max_share": float(args.diversity_max_share),
            "curated_min_count": int(target_curated_min),
            "curated_max_count": int(target_curated_max),
            "query_mode": args.query_mode,
            "llm_provider": args.llm_provider,
            "llm_model": args.llm_model,
            "embedding_provider": args.embedding_provider,
            "embedding_model": args.embedding_model,
            "embedding_backend": embedding_backend_name,
        },
        "grounding_files": {
            "xlsx": str(grounding_xlsx),
            "geojson": str(grounding_geojson),
            "docx": str(grounding_docx),
        },
    }
    report_path = out_dir / "pipeline_report.json"
    report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
    print(f"Report: {report_path}")
    print(textwrap.dedent(
        f"""
        Pipeline complete.
          Stage1 all:      {report['stage1_all_candidates']}
          Stage1 dedup:    {report['stage1_dedup_candidates']}
          Stage2 selected: {report['stage2_selected_by_metadata_count']}
          Stage3 cards:    {report['stage3_lightweight_card_count']}
          Stage6 enriched: {report['stage6_enriched_card_count']}
          Stage7 shortlist:{report['stage7_shortlist_count']}
          Final keep:      {report['final_keep_count']}
          Final review:    {report['final_review_count']}
        """
    ).strip())
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(
        description="GNEM document relevance pipeline: Tavily -> metadata scoring -> progressive document cards -> hybrid filtering -> classifier -> LLM judge -> credibility/diversity"
    )
    parser.add_argument("--queries-file", default=str(DEFAULT_QUERY_FILE), help="Path to query list")
    parser.add_argument("--input-metadata-xlsx", default="", help="Skip Tavily and start from existing metadata Excel")
    parser.add_argument("--golden-summary-file", default="", help="Path to text file containing golden summary")
    parser.add_argument("--api-key", default=os.environ.get("TAVILY_API_KEY", ""), help="Tavily API key")
    parser.add_argument("--output-dir", default="", help="Output directory for stage files")
    parser.add_argument(
        "--write-stage-artifacts",
        action="store_true",
        help="Write intermediate stage Excel/JSONL artifacts. Disabled by default so runs keep only final outputs.",
    )
    parser.add_argument(
        "--write-csv-exports",
        action="store_true",
        help="Also export final review-ready and registry CSV files. Disabled by default to reduce output clutter.",
    )
    parser.add_argument("--max-queries", type=int, default=1000, help="Max queries to run from query file")
    parser.add_argument(
        "--query-offset",
        type=int,
        default=0,
        help="Number of leading queries to skip before applying --max-queries (use 0, 250, 500, 750 for 250-query batches).",
    )
    parser.add_argument("--max-results", type=int, default=20, help="Tavily results per query")
    parser.add_argument("--search-depth", default="basic", choices=["basic", "advanced"], help="Tavily search depth")
    parser.add_argument(
        "--query-mode",
        default="hybrid",
        choices=["pdf_only", "web_only", "hybrid"],
        help="How to run each search query: force PDF, remove PDF bias for webpages, or run both variants",
    )
    parser.add_argument(
        "--disable-query-enhancement",
        action="store_true",
        help="Disable GNEM query expansion before sending query to Tavily",
    )
    parser.add_argument(
        "--no-raw-content",
        action="store_true",
        help="Disable Tavily raw content retrieval if you want lighter payloads",
    )
    parser.add_argument("--metadata-threshold", type=float, default=58.0, help="Stage-2 metadata score threshold")
    parser.add_argument(
        "--metadata-target-ratio",
        type=float,
        default=0.0,
        help="Optional Stage-2 keep ratio against Stage-1 dedup count (e.g., 0.5 for about half)",
    )
    parser.add_argument(
        "--disable-metadata-rule-gate",
        action="store_true",
        help="Only for calibration: do not require metadata 2-signal/question-specific rule in Stage 2",
    )
    parser.add_argument("--first-page-threshold", type=float, default=56.0, help="Stage-3 first-page score threshold")
    parser.add_argument("--final-threshold", type=float, default=60.0, help="Final combined score threshold")
    parser.add_argument(
        "--final-target-ratio",
        type=float,
        default=0.0,
        help="Optional final keep ratio against Stage-3 scored count (e.g., 0.5)",
    )
    parser.add_argument("--sample-size", type=int, default=0, help="Optional cap on Stage-2 rows for quick test (e.g., 50)")
    parser.add_argument("--timeout", type=int, default=90, help="HTTP timeout in seconds for downloads")
    parser.add_argument("--max-download-mb", type=float, default=80.0, help="Max single PDF size")
    parser.add_argument(
        "--llm-provider",
        default="ollama",
        choices=["none", "ollama", "openai"],
        help="LLM backend for shortlist-only document judging",
    )
    parser.add_argument(
        "--llm-model",
        default="qwen2.5:7b",
        help="Model name for selected LLM provider",
    )
    parser.add_argument(
        "--llm-base-url",
        default="http://localhost:11434",
        help="Base URL for LLM API (Ollama default: http://localhost:11434)",
    )
    parser.add_argument(
        "--llm-api-key",
        default=os.environ.get("OPENAI_API_KEY", ""),
        help="API key for provider=openai",
    )
    parser.add_argument(
        "--llm-timeout",
        type=int,
        default=120,
        help="Timeout in seconds for each LLM call",
    )
    parser.add_argument(
        "--llm-temperature",
        type=float,
        default=0.1,
        help="LLM temperature",
    )
    parser.add_argument(
        "--llm-max-text-chars",
        type=int,
        default=7000,
        help="Max document-card characters sent to the LLM judge",
    )
    parser.add_argument(
        "--llm-first-page-weight",
        type=float,
        default=0.75,
        help="Legacy compatibility flag; retained to avoid breaking older CLI invocations",
    )
    parser.add_argument(
        "--embedding-provider",
        default="ollama",
        choices=["none", "ollama", "openai"],
        help="Dense embedding backend used for chunk retrieval and reranking. Falls back to hashing when unavailable.",
    )
    parser.add_argument(
        "--embedding-model",
        default="nomic-embed-text",
        help="Embedding model name for the selected provider",
    )
    parser.add_argument(
        "--embedding-base-url",
        default="http://localhost:11434",
        help="Base URL for embedding API (Ollama default: http://localhost:11434)",
    )
    parser.add_argument(
        "--embedding-api-key",
        default=os.environ.get("OPENAI_API_KEY", ""),
        help="API key for provider=openai embeddings",
    )
    parser.add_argument(
        "--embedding-timeout",
        type=int,
        default=90,
        help="Timeout in seconds for embedding requests",
    )
    parser.add_argument(
        "--embedding-batch-size",
        type=int,
        default=12,
        help="Batch size for embedding requests",
    )
    parser.add_argument(
        "--local-pdf-dir",
        action="append",
        default=[],
        help="Optional local document dir(s) for Stage-2 acquisition without downloading (legacy flag name retained)",
    )
    parser.add_argument(
        "--local-text-dir",
        action="append",
        default=[],
        help="Optional local extracted-text dir(s) used to enrich document cards",
    )
    parser.add_argument(
        "--grounding-xlsx",
        default=str(DEFAULT_GROUNDING_XLSX),
        help="Grounding workbook used to build company/OEM dictionaries",
    )
    parser.add_argument(
        "--grounding-counties-geojson",
        default=str(DEFAULT_GROUNDING_GEOJSON),
        help="Grounding geojson used to build Georgia county dictionaries",
    )
    parser.add_argument(
        "--grounding-docx",
        default=str(DEFAULT_GROUNDING_DOCX),
        help="Grounding GNEM supply-chain vision docx",
    )
    parser.add_argument(
        "--heuristic-threshold",
        type=float,
        default=45.0,
        help="Minimum heuristic score for shortlist consideration",
    )
    parser.add_argument(
        "--embedding-threshold",
        type=float,
        default=65.0,
        help="Minimum semantic embedding score retained for backward-compatible reporting",
    )
    parser.add_argument(
        "--hybrid-threshold",
        type=float,
        default=None,
        help="Minimum hybrid score for shortlist consideration (defaults to --embedding-threshold when omitted)",
    )
    parser.add_argument(
        "--direct-usecase-threshold",
        type=float,
        default=0.70,
        help="Minimum direct use-case classifier score for shortlist consideration",
    )
    parser.add_argument(
        "--credibility-threshold",
        type=float,
        default=60.0,
        help="Minimum credibility score for keep decisions",
    )
    parser.add_argument(
        "--diversity-max-share",
        type=float,
        default=0.5,
        help="Maximum share of keep documents allowed from a single primary subtopic during the diversity pass",
    )
    parser.add_argument(
        "--document-max-chars",
        type=int,
        default=60000,
        help="Maximum number of document characters used for lightweight document cards before shortlist enrichment",
    )
    parser.add_argument(
        "--write-sqlite-registry",
        action="store_true",
        help="Also export the curated document/chunk registry to SQLite. Disabled by default so ingestion stays file-first.",
    )
    parser.add_argument(
        "--keep-stage-downloads",
        action="store_true",
        help="Keep the temporary stage2_downloads folder after the run. Disabled by default so final runs clean up temporary downloads.",
    )
    parser.add_argument(
        "--publish-ready-docs-dir",
        default=str(DEFAULT_EV_AUTOMATION_READY_DIR) if DEFAULT_EV_AUTOMATION_READY_DIR.parent.exists() else "",
        help="Optional publish target for curated Tavily documents, e.g. evAutomationUpdated/data/tavily ready documents.",
    )
    parser.add_argument(
        "--skip-ready-docs-publish",
        action="store_true",
        help="Skip publishing curated documents into the ready-docs handoff folder.",
    )
    parser.add_argument(
        "--curated-min-count",
        type=int,
        default=500,
        help="For large runs, promote additional strong candidates to review until at least this many curated documents remain.",
    )
    parser.add_argument(
        "--curated-max-count",
        type=int,
        default=1000,
        help="For large runs, stop recall-band promotions once this many curated documents have been reached.",
    )
    parser.add_argument(
        "--curated-band-min-stage1-count",
        type=int,
        default=10000,
        help="Only apply the curated recall band when Stage 1 has at least this many raw candidates.",
    )
    args = parser.parse_args()
    stage_artifacts_enabled = bool(args.write_stage_artifacts)

    if args.golden_summary_file:
        golden_summary = Path(args.golden_summary_file).read_text(encoding="utf-8").strip()
    else:
        golden_summary = GOLDEN_SUMMARY_DEFAULT

    llm_cfg: LLMConfig | None = None
    if args.llm_provider != "none":
        llm_base_url = args.llm_base_url
        if args.llm_provider == "openai" and llm_base_url.strip() == "http://localhost:11434":
            llm_base_url = "https://api.openai.com"
        llm_cfg = LLMConfig(
            provider=args.llm_provider,
            model=args.llm_model,
            base_url=llm_base_url,
            api_key=args.llm_api_key,
            timeout_sec=max(10, int(args.llm_timeout)),
            temperature=float(args.llm_temperature),
            max_text_chars=max(1000, int(args.llm_max_text_chars)),
            first_page_weight=max(0.0, min(1.0, float(args.llm_first_page_weight))),
        )
        if llm_cfg.provider == "openai" and not llm_cfg.api_key:
            print("Error: --llm-provider openai requires --llm-api-key or OPENAI_API_KEY.")
            return 1

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = Path(args.output_dir) if args.output_dir else DEFAULT_OUTPUTS_DIR / ts
    out_dir.mkdir(parents=True, exist_ok=True)
    print(f"Output dir: {out_dir.resolve()}")
    if llm_cfg:
        print(f"LLM mode: provider={llm_cfg.provider}, model={llm_cfg.model}, base_url={llm_cfg.base_url}")
    else:
        print("LLM mode: disabled (rule-based fallback only)")

    # Stage 1: gather and score metadata candidates
    if args.input_metadata_xlsx:
        print(f"Stage 1 source: metadata Excel -> {args.input_metadata_xlsx}")
        stage1_rows = metadata_rows_from_excel(
            input_xlsx=Path(args.input_metadata_xlsx),
            golden_summary=golden_summary,
        )
    else:
        if not args.api_key:
            print("Error: TAVILY_API_KEY is required unless --input-metadata-xlsx is used.")
            return 1
        q_path = resolve_existing_repo_path(
            args.queries_file,
            DEFAULT_QUERY_FILE,
            LEGACY_QUERY_FILE,
        ) or Path(args.queries_file)
        if not q_path.exists():
            print(f"Error: queries file not found: {q_path}")
            return 1
        queries = read_queries(q_path, max_queries=args.max_queries, query_offset=max(0, int(args.query_offset)))
        print(f"Loaded {len(queries)} queries from {q_path}")
        stage1_rows = tavily_search_rows(
            queries=queries,
            api_key=args.api_key,
            max_results=args.max_results,
            search_depth=args.search_depth,
            golden_summary=golden_summary,
            query_mode=args.query_mode,
            query_enhancement=not bool(args.disable_query_enhancement),
            include_raw_content=not bool(args.no_raw_content),
        )

    if not stage1_rows:
        print("No candidates found.")
        return 0

    stage1_df = pd.DataFrame(stage1_rows)
    stage1_df = stage1_df.sort_values(by=["Metadata_Score"], ascending=False).reset_index(drop=True)
    stage1_all_path = out_dir / "stage1_all_candidates.xlsx"
    maybe_write_excel(stage_artifacts_enabled, stage1_df, stage1_all_path)
    print(f"Stage 1 all candidates: {len(stage1_df)}" + (f" -> {stage1_all_path}" if stage_artifacts_enabled else ""))

    deduped_rows = dedupe_by_url_best_score(stage1_rows)
    stage1_dedup_df = pd.DataFrame(deduped_rows)
    stage1_dedup_path = out_dir / "stage1_dedup_best_by_url.xlsx"
    maybe_write_excel(stage_artifacts_enabled, stage1_dedup_df, stage1_dedup_path)
    print(f"Stage 1 dedup candidates: {len(stage1_dedup_df)}" + (f" -> {stage1_dedup_path}" if stage_artifacts_enabled else ""))

    return run_rag_filtering_pipeline(
        args=args,
        llm_cfg=llm_cfg,
        golden_summary=golden_summary,
        out_dir=out_dir,
        stage1_df=stage1_df,
        stage1_dedup_df=stage1_dedup_df,
    )

if __name__ == "__main__":
    raise SystemExit(main())
