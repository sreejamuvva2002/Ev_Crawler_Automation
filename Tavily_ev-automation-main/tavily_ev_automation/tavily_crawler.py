"""
Tavily Search Crawler - Outputs results to Excel in RAG Data Management format.
Downloads documents (PDFs + HTML pages + other file types).

Columns: Document_ID, Filename, File_Type, Category, Industry, Source, URL, Path,
         Size_MB, Pages, Date_Created, Date_Modified, Date_Accessed, Status,
         Language, Verified, Indexed, Access, Retention, Owner_Team, Score, Description.
"""

import os
import re
import mimetypes
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

import pandas as pd
import requests
from tavily import TavilyClient
from dotenv import load_dotenv
try:
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
except Exception:
    # Fallback for older openpyxl versions
    from openpyxl.utils.cell import ILLEGAL_CHARACTERS_RE
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:
    load_workbook = None
    Alignment = None
    Font = None
    PatternFill = None
    get_column_letter = None

def load_shared_env_files() -> None:
    repo_root = Path(__file__).resolve().parent.parent
    for env_path in [
        repo_root / ".env",
        repo_root / "evAutomationUpdated" / ".env",
    ]:
        if env_path.exists():
            load_dotenv(env_path, override=False)
    load_dotenv(override=False)


def disable_broken_local_proxies() -> None:
    broken_markers = ("127.0.0.1:9", "localhost:9")
    for key in ("HTTP_PROXY", "HTTPS_PROXY", "ALL_PROXY", "http_proxy", "https_proxy", "all_proxy"):
        value = (os.environ.get(key) or "").strip().lower()
        if value and any(marker in value for marker in broken_markers):
            os.environ.pop(key, None)


# Load .env files if present
load_shared_env_files()
disable_broken_local_proxies()

# Default config (customize as needed)
CONFIG = {
    "category": "GNEM Battery Supply Chain Explorer",
    "industry": "Georgia Battery Technology & Supply Chain",
    "base_path": "/data/raw/gnem/battery_supply_chain/georgia/",
    "owner_team": "GNEM Research Team",
    "max_results": 20,
    "search_depth": "advanced",
}

DEFAULT_OUTPUT_PATH = Path("outputs") / "crawler" / "georgia_ev_battery_manufacturing_suppliers.xlsx"



def get_file_type_from_url(url: str) -> str:
    """Infer file type from URL path as a best-effort fallback."""
    path = urlparse(url).path.lower()
    if path.endswith(".pdf"):
        return "PDF"
    if path.endswith((".htm", ".html")):
        return "HTML"
    if path.endswith((".doc", ".docx")):
        return "DOC"
    if path.endswith((".ppt", ".pptx")):
        return "PPT"
    if path.endswith((".xlsx", ".xls")):
        return "Excel"
    if path.endswith(".csv"):
        return "CSV"
    if path.endswith(".txt"):
        return "TXT"
    return "HTML"


def file_type_from_content_type(content_type: str) -> tuple[str, str | None]:
    """Map an HTTP Content-Type to (File_Type, extension)."""
    ct = (content_type or "").split(";", 1)[0].strip().lower()
    if not ct:
        return "", None

    if ct == "application/pdf":
        return "PDF", ".pdf"
    if ct in {"text/html", "application/xhtml+xml"}:
        return "HTML", ".html"
    if ct.startswith("text/plain"):
        return "TXT", ".txt"

    ext = mimetypes.guess_extension(ct)
    if ext == ".jpe":
        ext = ".jpg"

    if ct.startswith("image/"):
        return "Image", ext
    if ct.startswith("audio/"):
        return "Audio", ext
    if ct.startswith("video/"):
        return "Video", ext

    return ct.upper(), ext


def sanitize_filename(name: str) -> str:
    name = (name or "").strip()
    name = re.sub(r'[<>:"/\\|?*]', "_", name)
    name = re.sub(r"\s+", "_", name)
    name = re.sub(r"_+", "_", name)
    return name[:180] if len(name) > 180 else name


def url_to_filename(url: str, title: str | None) -> str:
    """Derive a filename from URL or title."""
    path = urlparse(url).path.strip("/")
    if path and "." in path.split("/")[-1]:
        return sanitize_filename(path.split("/")[-1])

    if title:
        safe = sanitize_filename(title)
        ext = ".pdf" if get_file_type_from_url(url) == "PDF" else ".html"
        return f"{safe}{ext}" if safe else f"page_{hash(url) % 10000}{ext}"

    name = urlparse(url).netloc.replace(".", "_") or "page"
    return f"{name}.html"


def domain_to_source(url: str) -> str:
    """Derive source/organization name from domain."""
    netloc = urlparse(url).netloc.lower()
    if not netloc:
        return "Unknown"
    name = re.sub(r"^www\.", "", netloc)
    name = re.sub(r"\.(com|org|gov|edu|net)$", "", name)
    return name.replace(".", " ").title()


def with_doc_id(filename: str, doc_id: str) -> str:
    """
    Ensure filename is stable + unique by appending _DOC_### before extension.
    """
    filename = sanitize_filename(filename)
    base, ext = os.path.splitext(filename)
    if base.endswith(f"_{doc_id}"):
        return filename
    return f"{base}_{doc_id}{ext}"


def update_row_from_existing(row: dict, full_path: str) -> None:
    try:
        size_mb = os.path.getsize(full_path) / (1024 * 1024)
    except OSError:
        size_mb = 0.0

    row["Path"] = os.path.dirname(full_path) + os.sep
    row["Filename"] = os.path.basename(full_path)
    row["Size_MB"] = round(size_mb, 2)
    row["Status"] = "Downloaded"

    # update File_Type from extension
    ext = os.path.splitext(full_path)[1].lower()
    if ext == ".pdf":
        row["File_Type"] = "PDF"
    elif ext in {".html", ".htm"}:
        row["File_Type"] = "HTML"


def run_search(
    query: str,
    api_key: str | None = None,
    max_results: int | None = None,
    search_depth: str | None = None,
    **config_overrides,
) -> list[dict]:
    """Run Tavily search and return list of row dicts for Excel."""
    api_key = api_key or os.environ.get("TAVILY_API_KEY")
    if not api_key:
        raise ValueError("Set TAVILY_API_KEY in environment or pass --api-key.")

    client = TavilyClient(api_key=api_key)
    cfg = {**CONFIG, **config_overrides}

    max_results = max_results or cfg["max_results"]
    depth = (search_depth or cfg.get("search_depth") or "basic").strip().lower()
    if depth not in {"basic", "advanced"}:
        depth = "basic"

    response = client.search(
        query=query,
        max_results=max_results,
        search_depth=depth,
    )

    rows: list[dict] = []
    today = datetime.now().strftime("%Y-%m-%d")

    for i, r in enumerate(response.get("results") or [], start=1):
        url = r.get("url", "")
        title = r.get("title") or ""
        content = (r.get("content") or "")[:500]
        score_raw = r.get("score") or 0.0
        score_int = min(100, max(0, round(score_raw * 100)))

        doc_id = f"DOC_{i:03d}"
        filename = with_doc_id(url_to_filename(url, title), doc_id)
        source = domain_to_source(url)

        rows.append(
            {
                "Document_ID": doc_id,
                "Filename": filename,
                "File_Type": get_file_type_from_url(url),
                "Category": cfg["category"],
                "Industry": cfg["industry"],
                "Source": source,
                "URL": url,
                "Path": "",
                "Size_MB": "",
                "Pages": "",
                "Date_Created": "",
                "Date_Modified": "",
                "Date_Accessed": today,
                "Status": "Discovered",
                "Language": "Unknown",
                "Verified": "Unknown",
                "Indexed": "No",
                "Access": "Public",
                "Retention": "TBD",
                "Owner_Team": cfg["owner_team"],
                "Score": score_int,
                "Description": content or title or url,
            }
        )

    return rows


def parse_content_disposition(cd: str | None) -> str | None:
    if not cd:
        return None
    m = re.search(r"filename\*=UTF-8''([^;]+)", cd, flags=re.IGNORECASE)
    if m:
        return sanitize_filename(requests.utils.unquote(m.group(1)))
    m = re.search(r'filename="([^"]+)"', cd, flags=re.IGNORECASE)
    if m:
        return sanitize_filename(m.group(1))
    m = re.search(r"filename=([^;]+)", cd, flags=re.IGNORECASE)
    if m:
        return sanitize_filename(m.group(1).strip().strip('"'))
    return None


def download_url(
    url: str,
    save_path: str,
    timeout: int = 90,
    max_download_mb: float = 50.0,
) -> tuple[bool, float, str | None, str | None, str | None]:
    """
    Download URL to save_path.
    Returns: (success, size_mb, content_type, content_disposition, final_url)
    """
    parsed = urlparse(url)
    referer = f"{parsed.scheme}://{parsed.netloc}/" if parsed.scheme and parsed.netloc else ""

    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,application/pdf;q=0.8,*/*;q=0.7",
        "Accept-Language": "en-US,en;q=0.9",
        "Referer": referer,
        "DNT": "1",
        "Connection": "keep-alive",
    }

    try:
        with requests.get(
            url, headers=headers, timeout=timeout, stream=True, allow_redirects=True
        ) as r:
            r.raise_for_status()
            content_type = (r.headers.get("Content-Type") or "").strip()
            content_disp = (r.headers.get("Content-Disposition") or "").strip() or None
            final_url = r.url or url

            cl = r.headers.get("Content-Length")
            if cl and cl.isdigit():
                size_mb = int(cl) / (1024 * 1024)
                if size_mb > max_download_mb:
                    print(f"  Skipping (>{max_download_mb} MB): {final_url}")
                    return False, 0.0, content_type, content_disp, final_url

            bytes_written = 0
            max_bytes = int(max_download_mb * 1024 * 1024)

            with open(save_path, "wb") as f:
                for chunk in r.iter_content(chunk_size=1024 * 128):
                    if not chunk:
                        continue
                    f.write(chunk)
                    bytes_written += len(chunk)
                    if bytes_written > max_bytes:
                        print(f"  Aborting (>{max_download_mb} MB): {final_url}")
                        f.close()
                        try:
                            os.remove(save_path)
                        except OSError:
                            pass
                        return False, 0.0, content_type, content_disp, final_url

            size_mb = bytes_written / (1024 * 1024)
            return True, round(size_mb, 2), content_type, content_disp, final_url

    except (requests.RequestException, OSError) as e:
        print(f"  Download failed: {url} -> {e}")
        return False, 0.0, None, None, None


def download_documents(
    rows: list[dict],
    download_dir: str,
    download_mode: str = "all",
    timeout: int = 90,
    max_download_mb: float = 50.0,
) -> None:
    """
    Download results to download_dir.

    download_mode:
      - "all": download PDFs + HTML + anything else
      - "pdf": download only PDFs
      - "html": download only HTML pages

    Skips already-downloaded files (stable naming with _DOC_###).
    Updates Path, Size_MB, Filename, File_Type in each row on success.
    """
    os.makedirs(download_dir, exist_ok=True)
    mode = (download_mode or "all").strip().lower()
    if mode not in {"all", "pdf", "html"}:
        mode = "all"

    for row in rows:
        url = row.get("URL", "")
        if not url:
            continue

        ft_guess = row.get("File_Type") or get_file_type_from_url(url)
        if mode == "pdf" and ft_guess != "PDF":
            continue
        if mode == "html" and ft_guess != "HTML":
            continue

        doc_id = row.get("Document_ID", "DOC_000")
        filename = row.get("Filename") or url_to_filename(url, None)
        filename = with_doc_id(filename, doc_id)
        if not filename:
            filename = f"page_{hash(url) % 100000}_{doc_id}.html"

        save_path = os.path.join(download_dir, filename)

        # Skip if already downloaded
        if os.path.exists(save_path):
            print(f"Skipping existing: {os.path.basename(save_path)}")
            update_row_from_existing(row, save_path)
            continue

        print(f"Downloading: {os.path.basename(save_path)}")

        tmp_path = save_path + ".part"
        ok, size_mb, content_type, content_disp, final_url = download_url(
            url, tmp_path, timeout=timeout, max_download_mb=max_download_mb
        )

        if not ok:
            row["Status"] = "Failed"
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except OSError:
                pass
            continue

        detected_type, detected_ext = file_type_from_content_type(content_type or "")
        if final_url:
            row["URL"] = final_url

        cd_name = parse_content_disposition(content_disp)

        # Start from content-disposition name if available, else current filename
        base_name = cd_name or os.path.basename(save_path)
        base_name = sanitize_filename(base_name)
        base, ext = os.path.splitext(base_name)

        if detected_ext and ext.lower() != detected_ext.lower():
            if ext.lower() in {"", ".htm", ".html", ".pdf", ".bin"}:
                base_name = f"{base}{detected_ext}"

        # Ensure stable doc id suffix
        final_name = with_doc_id(base_name, doc_id)
        final_path = os.path.join(download_dir, final_name)

        # If collision (rare), add a counter
        if os.path.exists(final_path):
            b2, e2 = os.path.splitext(final_name)
            k = 2
            while os.path.exists(os.path.join(download_dir, f"{b2}_{k}{e2}")):
                k += 1
            final_path = os.path.join(download_dir, f"{b2}_{k}{e2}")

        try:
            os.replace(tmp_path, final_path)
        except OSError:
            os.replace(tmp_path, save_path)
            final_path = save_path

        row["Path"] = os.path.dirname(final_path) + os.sep
        row["Size_MB"] = size_mb
        row["Filename"] = os.path.basename(final_path)

        if detected_type:
            if detected_type == "TEXT/HTML":
                row["File_Type"] = "HTML"
            elif detected_type == "APPLICATION/PDF":
                row["File_Type"] = "PDF"
            elif detected_type in {
                "HTML",
                "PDF",
                "DOC",
                "PPT",
                "Excel",
                "CSV",
                "TXT",
                "Image",
                "Audio",
                "Video",
            }:
                row["File_Type"] = detected_type
            else:
                row["File_Type"] = ft_guess

        row["Status"] = "Downloaded"


def export_to_excel(rows: list[dict], output_path: str) -> None:
    """Write rows to Excel with correct column order."""
    columns = [
        "Document_ID",
        "Filename",
        "File_Type",
        "Category",
        "Industry",
        "Source",
        "URL",
        "Path",
        "Size_MB",
        "Pages",
        "Date_Created",
        "Date_Modified",
        "Date_Accessed",
        "Status",
        "Language",
        "Verified",
        "Indexed",
        "Access",
        "Retention",
        "Owner_Team",
        "Score",
        "Description",
    ]
    cleaned_rows: list[dict] = []
    for row in rows:
        cleaned: dict = {}
        for key, value in row.items():
            if isinstance(value, str):
                cleaned[key] = ILLEGAL_CHARACTERS_RE.sub("", value)
            else:
                cleaned[key] = value
        cleaned_rows.append(cleaned)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(cleaned_rows, columns=columns)
    df.to_excel(output, index=False, engine="openpyxl")
    if load_workbook is not None and Font is not None and PatternFill is not None and Alignment is not None and get_column_letter is not None:
        workbook = load_workbook(output)
        worksheet = workbook.active
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_idx in range(1, worksheet.max_column + 1):
            column_letter = get_column_letter(column_idx)
            max_len = max(len(str(cell.value or "")) for cell in worksheet[column_letter])
            worksheet.column_dimensions[column_letter].width = min(max(12, max_len + 2), 42)
        workbook.save(output)
    print(f"Exported {len(rows)} rows to {output}")


def main() -> int:
    import argparse

    parser = argparse.ArgumentParser(
        description="Tavily crawler -> Excel (RAG format) + downloads"
    )
    parser.add_argument(
        "query",
        nargs="?",
        default="Georgia EV battery manufacturing suppliers",
        help="Search query for Tavily",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=str(DEFAULT_OUTPUT_PATH),
        help="Output Excel file path",
    )
    parser.add_argument(
        "-n",
        "--max-results",
        type=int,
        default=20,
        help="Max search results (1-20)",
    )
    parser.add_argument(
        "--search-depth",
        default=CONFIG.get("search_depth", "basic"),
        choices=["basic", "advanced"],
        help="Tavily search depth",
    )
    parser.add_argument(
        "--api-key",
        default=os.environ.get("TAVILY_API_KEY", ""),
        help="Tavily API key (or set TAVILY_API_KEY)",
    )
    parser.add_argument(
        "-d",
        "--download-dir",
        default=None,
        metavar="DIR",
        help="Download results to this folder (PDFs + HTML + other types)",
    )
    parser.add_argument(
        "--download-mode",
        default="all",
        choices=["all", "pdf", "html"],
        help="What to download when --download-dir is set",
    )
    parser.add_argument(
        "--max-download-mb",
        type=float,
        default=50.0,
        help="Skip/abort any single download larger than this size (MB)",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=90,
        help="HTTP timeout in seconds",
    )

    args = parser.parse_args()

    if not args.api_key:
        print("Error: Set TAVILY_API_KEY in .env or pass --api-key")
        return 1

    rows = run_search(
        args.query,
        api_key=args.api_key,
        max_results=args.max_results,
        search_depth=args.search_depth,
    )

    if not rows:
        print("No results returned.")
        return 0

    if args.download_dir:
        print(f"Downloading ({args.download_mode}) to: {args.download_dir}")
        download_documents(
            rows,
            args.download_dir,
            download_mode=args.download_mode,
            timeout=args.timeout,
            max_download_mb=args.max_download_mb,
        )

    export_to_excel(rows, args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
